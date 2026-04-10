import streamlit as st
import pandas as pd
import re
import unicodedata
import io
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import (Mail, Attachment, FileContent,
                                        FileName, FileType, Disposition, Cc)
    SENDGRID_OK = True
except ImportError:
    SENDGRID_OK = False

st.set_page_config(
    page_title="CS Rehagro | Engajamento",
    page_icon="🌱",
    layout="wide",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Clash+Display:wght@500;600;700&family=Cabinet+Grotesk:wght@400;500;700;800&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Outfit:wght@300;400;500;600;700&display=swap');

:root {
  --verde:   #0F3D20;
  --verde2:  #1B5E35;
  --verde3:  #2E7D32;
  --ouro:    #C8A951;
  --ouro2:   #E8C96A;
  --creme:   #F4F0E6;
  --cinza:   #EDEAE0;
  --texto:   #111111;
  --sub:     #5A5A4A;
  --branco:  #FFFFFF;
}

*, *::before, *::after { box-sizing: border-box; }
html, body, [class*="css"] {
    font-family: 'Outfit', sans-serif !important;
    background: var(--creme) !important;
    color: var(--texto) !important;
}
.stApp { background: var(--creme) !important; }
#MainMenu, footer, header { visibility: hidden; }
[data-testid="stAppViewContainer"] { padding-top: 0 !important; }
[data-testid="block-container"] { padding-top: 0 !important; }

/* ── HERO ─────────────────────────────── */
.rh-hero {
    background: var(--verde);
    padding: 18px 40px 20px;
    margin: -1rem -1rem 0 -1rem;
    position: relative;
    overflow: hidden;
    border-bottom: 3px solid var(--ouro);
}
.rh-hero-bg {
    position: absolute; inset: 0;
    background:
        radial-gradient(ellipse 80% 60% at 110% 50%, rgba(200,169,81,0.18) 0%, transparent 60%),
        radial-gradient(ellipse 40% 80% at -10% 50%, rgba(46,125,50,0.3) 0%, transparent 60%);
}
.rh-hero-grid {
    position: absolute; inset: 0; opacity: 0.04;
    background-image: linear-gradient(var(--ouro) 1px, transparent 1px),
                      linear-gradient(90deg, var(--ouro) 1px, transparent 1px);
    background-size: 40px 40px;
}
.rh-hero-inner { position: relative; z-index: 1; }
.rh-eyebrow {
    font-family: 'Outfit', sans-serif;
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 4px;
    text-transform: uppercase;
    color: var(--ouro);
    margin: 0 0 6px 0;
    display: flex;
    align-items: center;
    gap: 10px;
}
.rh-eyebrow::before {
    content: '';
    display: inline-block;
    width: 32px; height: 2px;
    background: var(--ouro);
}
.rh-hero-title {
    font-family: 'Bebas Neue', sans-serif;
    font-size: 2.2rem;
    line-height: 1;
    color: var(--branco);
    margin: 0 0 4px 0;
    letter-spacing: 2px;
    white-space: nowrap;
}
.rh-hero-title span { color: var(--ouro); }
.rh-hero-sub {
    font-size: 0.85rem;
    color: rgba(255,255,255,0.6);
    font-weight: 300;
    margin: 4px 0 0 0;
    line-height: 1.4;
    white-space: nowrap;
}
.rh-hero-pills {
    display: flex; gap: 6px; margin-top: 10px; flex-wrap: nowrap;
}
.rh-pill {
    background: rgba(200,169,81,0.15);
    border: 1px solid rgba(200,169,81,0.35);
    color: var(--ouro2);
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 1px;
    text-transform: uppercase;
    padding: 5px 14px;
    border-radius: 100px;
}

/* ── BODY WRAPPER ─────────────────────── */
.rh-body { padding: 40px 0 0 0; }

/* ── SECTION LABEL ───────────────────── */
.rh-section {
    font-family: 'Outfit', sans-serif;
    font-size: 0.65rem;
    font-weight: 700;
    letter-spacing: 3.5px;
    text-transform: uppercase;
    color: var(--ouro);
    margin: 0 0 20px 0;
    display: flex;
    align-items: center;
    gap: 10px;
}
.rh-section::after {
    content: '';
    flex: 1;
    height: 1px;
    background: linear-gradient(90deg, rgba(200,169,81,0.4), transparent);
}

/* ── DASHBOARD CARDS ─────────────────── */
.rh-dash-card {
    background: var(--branco);
    border-radius: 0;
    padding: 24px 28px;
    margin-bottom: 2px;
    border-left: 3px solid transparent;
    transition: border-color 0.2s;
    position: relative;
}
.rh-dash-card:first-of-type { border-radius: 12px 12px 0 0; }
.rh-dash-card:last-of-type  { border-radius: 0 0 12px 12px; margin-bottom: 0; }
.rh-dash-card:hover { border-left-color: var(--ouro); }

.rh-dash-num {
    font-family: 'Bebas Neue', sans-serif !important;
    font-size: 1.1rem !important;
    letter-spacing: 3px;
    color: var(--ouro);
    margin-bottom: 4px;
}
.rh-dash-title {
    font-size: 1.05rem;
    font-weight: 700;
    color: var(--verde);
    margin-bottom: 12px;
    line-height: 1.2;
}
.rh-dash-desc {
    font-size: 0.88rem;
    color: var(--sub);
    line-height: 1.7;
}
.rh-tag {
    display: inline-block;
    background: #EDF7EE;
    color: var(--verde3);
    border: 1px solid rgba(46,125,50,0.18);
    font-size: 0.75rem;
    font-weight: 500;
    padding: 3px 10px;
    border-radius: 4px;
    margin: 2px 2px 2px 0;
}
.rh-note {
    font-size: 0.76rem;
    color: var(--ouro);
    margin-top: 10px;
    font-style: italic;
}
.rh-opt-badge {
    display: inline-block;
    background: rgba(200,169,81,0.12);
    border: 1px solid rgba(200,169,81,0.3);
    color: var(--ouro);
    font-size: 0.65rem;
    font-weight: 700;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    padding: 2px 8px;
    border-radius: 4px;
    margin-bottom: 6px;
}

/* ── UPLOAD LABELS ───────────────────── */
[data-testid="stFileUploader"] label,
[data-testid="stFileUploader"] > div > label {
    font-size: 1rem !important;
    font-weight: 600 !important;
    color: var(--verde) !important;
    margin-bottom: 6px !important;
}
[data-testid="stFileUploader"] {
    background: var(--branco) !important;
    border-radius: 10px !important;
    border: 1.5px dashed rgba(15,61,32,0.2) !important;
    padding: 6px 10px !important;
    margin-bottom: 12px !important;
}
[data-testid="stFileUploader"]:focus-within {
    border-color: var(--ouro) !important;
    box-shadow: 0 0 0 3px rgba(200,169,81,0.15) !important;
}

/* ── TEXT INPUT ──────────────────────── */
[data-testid="stTextInput"] label {
    font-size: 1rem !important;
    font-weight: 600 !important;
    color: var(--verde) !important;
}
[data-testid="stTextInput"] input {
    border-radius: 8px !important;
    border: 1.5px solid rgba(15,61,32,0.2) !important;
    font-size: 0.95rem !important;
}

/* ── BUTTON ──────────────────────────── */
.stButton > button[kind="primary"] {
    background: var(--verde) !important;
    color: var(--branco) !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'Outfit', sans-serif !important;
    font-size: 1rem !important;
    font-weight: 700 !important;
    letter-spacing: 0.5px !important;
    padding: 14px 28px !important;
    box-shadow: 0 4px 20px rgba(15,61,32,0.25) !important;
    transition: all 0.2s !important;
    position: relative !important;
    overflow: hidden !important;
}
.stButton > button[kind="primary"]::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 3px;
    background: var(--ouro);
}
.stButton > button[kind="primary"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 28px rgba(15,61,32,0.35) !important;
}

/* ── DOWNLOAD BUTTON ─────────────────── */
.stDownloadButton > button {
    background: transparent !important;
    color: var(--verde) !important;
    border: 2px solid var(--verde) !important;
    border-radius: 8px !important;
    font-family: 'Outfit', sans-serif !important;
    font-size: 0.95rem !important;
    font-weight: 600 !important;
    padding: 12px 24px !important;
}
.stDownloadButton > button:hover {
    background: var(--verde) !important;
    color: var(--branco) !important;
}

/* ── MÉTRICAS ────────────────────────── */
.rh-metrics {
    display: grid;
    grid-template-columns: repeat(4,1fr);
    gap: 10px;
    margin: 24px 0;
}
.rh-metric {
    background: var(--branco);
    border-radius: 10px;
    padding: 20px 14px;
    text-align: center;
    border: 1px solid var(--cinza);
    position: relative;
    overflow: hidden;
}
.rh-metric::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 3px;
}
.rh-metric.m-total::after  { background: var(--verde3); }
.rh-metric.m-crit::after   { background: #C62828; }
.rh-metric.m-atenc::after  { background: #E65100; }
.rh-metric.m-mon::after    { background: #F57F17; }
.rh-metric-num {
    font-family: 'Bebas Neue', sans-serif;
    font-size: 3rem;
    line-height: 1;
    margin-bottom: 4px;
}
.rh-metric-label {
    font-size: 0.72rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    color: var(--sub);
}

/* ── DIVIDER ─────────────────────────── */
.rh-divider {
    height: 1px;
    background: linear-gradient(90deg, var(--ouro), transparent);
    margin: 24px 0;
    opacity: 0.35;
}

/* ── EXPANDER ────────────────────────── */
[data-testid="stExpander"] {
    border: 1px solid var(--cinza) !important;
    border-radius: 8px !important;
    background: var(--branco) !important;
}

/* ── DATAFRAME ───────────────────────── */
[data-testid="stDataFrame"] {
    border-radius: 10px !important;
    overflow: hidden !important;
    border: 1px solid var(--cinza) !important;
}

/* ── SUCCESS / INFO / WARNING ────────── */
[data-testid="stAlert"] {
    border-radius: 8px !important;
    font-size: 0.9rem !important;
}

/* ── FOOTER ──────────────────────────── */
.rh-footer {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 16px;
    padding: 32px 0 16px;
    border-top: 1px solid var(--cinza);
    margin-top: 48px;
    color: var(--sub);
    font-size: 0.8rem;
}
.rh-footer-dot {
    width: 4px; height: 4px;
    border-radius: 50%;
    background: var(--ouro);
    display: inline-block;
}
</style>
""", unsafe_allow_html=True)

# ── Configurações de e-mail ──────────────────────────────
DESTINATARIOS_FIXOS = ["rafael.ferraz@rehagro.edu.br"]
REMETENTE = "rafael.ferraz@rehagro.edu.br"


# ── Funções de processamento ─────────────────────────────


import unicodedata, re

# ── Helper global ────────────────────────────────────────────
def _norm(n):
    n = str(n).strip().lower()
    n = unicodedata.normalize('NFD', n)
    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
    n = re.sub(r'\s+', ' ', n)
    return n

_MESES = {'janeiro':1,'fevereiro':2,'março':3,'abril':4,'maio':5,'junho':6,
          'julho':7,'agosto':8,'setembro':9,'outubro':10,'novembro':11,'dezembro':12}

def _parse_freq_date(s):
    m = re.match(r'(\d{2}/\d{2}/\d{4})', str(s))
    return pd.to_datetime(m.group(1), dayfirst=True) if m else pd.NaT

def _montar_data_nps(row):
    try:
        ano = int(row['Data da aula - Ano'])
        mes = _MESES.get(str(row['Data da aula - Mês']).strip().lower())
        dia = int(row['Data da aula - Dia'])
        if mes: return pd.Timestamp(year=ano, month=mes, day=dia)
    except: pass
    return pd.NaT



import unicodedata, re

# ── Helper global ────────────────────────────────────────────
def _norm(n):
    n = str(n).strip().lower()
    n = unicodedata.normalize('NFD', n)
    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
    n = re.sub(r'\s+', ' ', n)
    return n

_MESES = {'janeiro':1,'fevereiro':2,'março':3,'abril':4,'maio':5,'junho':6,
          'julho':7,'agosto':8,'setembro':9,'outubro':10,'novembro':11,'dezembro':12}

def _parse_freq_date(s):
    m = re.match(r'(\d{2}/\d{2}/\d{4})', str(s))
    return pd.to_datetime(m.group(1), dayfirst=True) if m else pd.NaT

def _montar_data_nps(row):
    try:
        ano = int(row['Data da aula - Ano'])
        mes = _MESES.get(str(row['Data da aula - Mês']).strip().lower())
        dia = int(row['Data da aula - Dia'])
        if mes: return pd.Timestamp(year=ano, month=mes, day=dia)
    except: pass
    return pd.NaT


def carregar_canvas(arquivo):
    df = pd.read_excel(arquivo, skiprows=2)
    df.columns = df.columns.str.strip()
    col_nome   = next((c for c in df.columns if 'NOME' in c.upper()), None)
    col_email  = next((c for c in df.columns if 'E-MAIL' in c.upper() or 'EMAIL' in c.upper()), None)
    col_dias   = next((c for c in df.columns if 'DIAS' in c.upper()), None)
    col_curso  = next((c for c in df.columns if c.upper() == 'CURSO'), None)
    col_turma  = next((c for c in df.columns if c.upper() == 'TURMA'), None)
    col_funcao = next((c for c in df.columns if 'FUN' in c.upper() and 'DISCIPLINA' in c.upper()), None)
    if not col_nome or not col_dias:
        raise ValueError("Arquivo Canvas: colunas não encontradas.")
    if col_funcao:
        df = df[df[col_funcao].astype(str).str.upper().str.contains('ALUNO', na=False)].copy()
    df['_key']  = df[col_nome].apply(_norm)
    df['_dias'] = pd.to_numeric(df[col_dias], errors='coerce')
    cols = {'_key':'_key', col_nome:'Nome', '_dias':'Dias sem acesso'}
    if col_email: cols[col_email] = 'Email'
    if col_curso: cols[col_curso] = 'Curso'
    if col_turma: cols[col_turma] = 'Turma'
    res = df[df['_dias'] > 20][list(cols.keys())].copy().rename(columns=cols)
    if 'Email' not in res.columns: res['Email'] = ''
    if 'Curso' not in res.columns: res['Curso'] = ''
    if 'Turma' not in res.columns: res['Turma'] = ''
    return res[['_key','Nome','Email','Curso','Turma','Dias sem acesso']]


def carregar_frequencia(arquivo):
    df = pd.read_excel(arquivo, skiprows=2)
    df.columns = df.columns.str.strip()
    col_aluno  = next((c for c in df.columns if c.upper() == 'ALUNO'), None)
    col_status = next((c for c in df.columns if 'STATUS' in c.upper() or 'PRESEN' in c.upper()), None)
    col_data   = next((c for c in df.columns if 'DATA' in c.upper() or 'PARTE' in c.upper()), None)
    col_turma  = next((c for c in df.columns if c.upper() == 'TURMA'), None)
    if not col_aluno or not col_status:
        raise ValueError("Arquivo Frequência: colunas não encontradas.")
    df['_key'] = df[col_aluno].apply(_norm)
    df['_data'] = df[col_data].apply(_parse_freq_date) if col_data else pd.NaT
    # Desistentes em qualquer momento → excluir de TODOS os alertas
    desistentes_keys = set(
        df[df[col_status].astype(str).str.upper() == 'DESISTENTE']['_key'].unique()
    )
    # Turma por aluno (para relatório)
    turma_map = {}
    if col_turma:
        for _, row in df.iterrows():
            k = row['_key']
            if k not in turma_map and pd.notna(row[col_turma]):
                turma_map[k] = str(row[col_turma]).strip()
    # Apenas ativos
    df_a = df[~df['_key'].isin(desistentes_keys) & (df[col_status].astype(str).str.strip() != '-')].copy()
    df_a = df_a.sort_values(['_key', '_data'])
    ausentes = []
    for nome, g in df_a.groupby('_key'):
        if len(g) < 2: continue
        last2 = g.tail(2)
        status_last2 = last2[col_status].astype(str).str.upper().tolist()
        if status_last2 == ['AUSENTE', 'AUSENTE']:
            datas = last2[col_data].tolist() if col_data else ['N/D','N/D']
            ausentes.append({'_key': nome, 'Ultimas_2_aulas': f"{datas[0]} e {datas[1]}"})
    df_ausentes = pd.DataFrame(ausentes) if ausentes else pd.DataFrame(columns=['_key','Ultimas_2_aulas'])
    return df_ausentes, desistentes_keys, turma_map, df_a


def carregar_comentarios(arquivo):
    df = pd.read_excel(arquivo, skiprows=2)
    df.columns = df.columns.str.strip()
    col_aluno  = next((c for c in df.columns if c.upper() == 'ALUNO'), None)
    col_data   = next((c for c in df.columns if 'DATA DA AULA' in c.upper() or c.upper() == 'DATA DA AULA'), None)
    col_topico = next((c for c in df.columns if 'TÓPICO' in c.upper() or 'TOPICO' in c.upper()), None)
    col_prof   = next((c for c in df.columns if 'PROFESSOR' in c.upper()), None)
    col_resp   = next((c for c in df.columns if 'RESPOSTA' in c.upper()), None)
    if not col_aluno or not col_resp:
        raise ValueError("Arquivo Comentários: colunas não encontradas.")
    df = df[df[col_resp].notna() & (df[col_resp].astype(str).str.strip() != '')].copy()
    df['_key']    = df[col_aluno].apply(_norm)
    df['_data']   = pd.to_datetime(df[col_data], errors='coerce').dt.normalize() if col_data else pd.NaT
    df['_topico'] = df[col_topico].astype(str).str.strip().str.lower() if col_topico else ''
    df['_prof']   = df[col_prof].astype(str).str.strip() if col_prof else ''
    df['_topico_orig'] = df[col_topico].astype(str).str.strip() if col_topico else ''
    df['_resp']   = df[col_resp].astype(str).str.strip()
    return df[['_key','_data','_topico','_topico_orig','_prof','_resp']]


def carregar_nps(arquivo, desistentes_keys=None):
    df = pd.read_excel(arquivo, skiprows=2)
    df.columns = df.columns.str.strip()
    col_aluno  = next((c for c in df.columns if c.upper() == 'ALUNO'), None)
    col_nps    = next((c for c in df.columns if 'NPS' in c.upper() and 'REA' in c.upper()), None)
    col_topico = next((c for c in df.columns if 'TÓPICO' in c.upper() or 'TOPICO' in c.upper()), None)
    col_prof   = next((c for c in df.columns if 'PROFESSOR' in c.upper()), None)
    if not col_aluno or not col_nps:
        raise ValueError("Arquivo NPS: colunas não encontradas.")
    # Data correta = Ano + Mês + Dia (colunas A, B, C)
    df['_data'] = df.apply(_montar_data_nps, axis=1)
    df['_key']  = df[col_aluno].apply(_norm)
    df['_topico']      = df[col_topico].astype(str).str.strip().str.lower() if col_topico else ''
    df['_topico_orig'] = df[col_topico].astype(str).str.strip() if col_topico else ''
    df['_prof']        = df[col_prof].astype(str).str.strip() if col_prof else ''
    df['_nps']         = pd.to_numeric(df[col_nps], errors='coerce')
    if desistentes_keys:
        df = df[~df['_key'].isin(desistentes_keys)].copy()
    df = df.sort_values(['_key','_data'])
    return df[['_key','_data','_topico','_topico_orig','_prof','_nps']]


def gerar_alertas_nps(df_nps, df_coment, df_freq_ativo):
    """
    Gera alertas NPS por aluno com critérios A-F.
    Retorna dict: {_key: [{'texto':..., 'acao':..., 'topico':..., 'professor':...}]}
    """
    alertas_por_aluno = {}

    def fmt_data(d):
        try: return pd.Timestamp(d).strftime('%d/%m/%Y')
        except: return str(d)

    def add(key, texto, acao, topico='', professor='', comentario=''):
        alertas_por_aluno.setdefault(key, []).append(
            {'texto': texto, 'acao': acao, 'topico': topico, 'professor': professor, 'comentario': comentario}
        )

    # Mapa de comentários: (key, data, topico) → {prof, resp}
    coment_map = {}
    if df_coment is not None and not df_coment.empty:
        for _, row in df_coment.iterrows():
            k = (row['_key'], row['_data'], row['_topico'])
            coment_map[k] = {'prof': row['_prof'], 'resp': row['_resp'],
                             'topico_orig': row['_topico_orig']}

    # Aulas de frequência por aluno (para critérios B e F)
    freq_por_aluno = {}
    if df_freq_ativo is not None and not df_freq_ativo.empty:
        for nome, g in df_freq_ativo.groupby('_key'):
            freq_por_aluno[nome] = g.sort_values('_data')

    todos_keys = set(df_nps['_key'].unique()) | set(df_coment['_key'].unique() if df_coment is not None else [])

    for key in todos_keys:
        nps_aluno    = df_nps[df_nps['_key'] == key].sort_values('_data')
        coment_aluno = df_coment[df_coment['_key'] == key] if df_coment is not None else pd.DataFrame()
        freq_aluno   = freq_por_aluno.get(key, pd.DataFrame())

        alertas_gerados = set()  # evitar duplicatas por aula

        # ── A: Detrator nos últimos 2 encontros avaliados ──────────
        respondidas = nps_aluno[nps_aluno['_nps'].notna()]
        ult2_resp   = respondidas.tail(2)
        if len(ult2_resp) == 2 and all(ult2_resp['_nps'] < 0):
            detalhes = []
            for _, r in ult2_resp.iterrows():
                detalhes.append(f"{fmt_data(r['_data'])} · {r['_topico_orig']} · Prof. {r['_prof']}")
                alertas_gerados.add((r['_data'], r['_topico']))
            topicos  = ' · '.join([r['_topico_orig'] for _,r in ult2_resp.iterrows()])
            profs    = ' · '.join([r['_prof'] for _,r in ult2_resp.iterrows()])
            add(key,
                f"Detrator nos últimos 2 encontros: {' || '.join(detalhes)}",
                "Retomar feedback negativo da avaliação",
                topicos, profs)

        # ── B: Presente nas últimas 2 aulas sem avaliar nenhuma ────
        if not freq_aluno.empty and len(freq_aluno) >= 2:
            ult2_freq = freq_aluno.tail(2)
            status_ult2 = ult2_freq['Primeiro StatusPresenca'].astype(str).str.upper().tolist() \
                if 'Primeiro StatusPresenca' in ult2_freq.columns else []
            if status_ult2 == ['PRESENTE', 'PRESENTE']:
                datas_freq = ult2_freq['_data'].tolist()
                nps_nessas = nps_aluno[nps_aluno['_data'].isin(datas_freq) & nps_aluno['_nps'].notna()]
                if nps_nessas.empty:
                    datas_str = ' e '.join([fmt_data(d) for d in datas_freq])
                    add(key,
                        f"Presente nas últimas 2 aulas sem avaliar ({datas_str})",
                        "Incentivar participação nas avaliações de aula")

        # ── F: Detrator na penúltima + ausente na última ────────────
        if not freq_aluno.empty and len(freq_aluno) >= 2:
            ult   = freq_aluno.iloc[-1]
            penul = freq_aluno.iloc[-2]
            status_cols = [c for c in freq_aluno.columns if 'STATUS' in c.upper() or 'PRESEN' in c.upper()]
            if status_cols:
                sc = status_cols[0]
                if str(ult[sc]).upper() == 'AUSENTE':
                    data_penul  = penul['_data']
                    nps_penul   = nps_aluno[nps_aluno['_data'] == data_penul]
                    det_penul   = nps_penul[nps_penul['_nps'] < 0]
                    if not det_penul.empty:
                        r = det_penul.iloc[0]
                        chave_aula = (r['_data'], r['_topico'])
                        if chave_aula not in alertas_gerados:
                            add(key,
                                f"Detrator na penúltima aula ({fmt_data(r['_data'])} · {r['_topico_orig']} · Prof. {r['_prof']}) e ausente na última ({fmt_data(ult['_data'])})",
                                "Retomar feedback e verificar ausência",
                                r['_topico_orig'], r['_prof'])
                            alertas_gerados.add(chave_aula)

        # ── C/D/E: Comentários ──────────────────────────────────────
        if not coment_aluno.empty:
            for _, crow in coment_aluno.iterrows():
                c_data   = crow['_data']
                c_topico = crow['_topico']
                c_key    = (c_data, c_topico)

                # Verificar se há NPS negativo na mesma aula
                nps_mesma = nps_aluno[(nps_aluno['_data'] == c_data) &
                                      (nps_aluno['_topico'] == c_topico) &
                                      (nps_aluno['_nps'] < 0)]

                if not nps_mesma.empty and c_key not in alertas_gerados:
                    # D: Detrator + comentário na mesma aula → 1 alerta unificado
                    r = nps_mesma.iloc[0]
                    add(key,
                        f"Escreveu comentário em {fmt_data(c_data)} (com avaliação negativa)",
                        "Retomar feedback negativo e comentário da avaliação",
                        crow['_topico_orig'], crow['_prof'],
                        comentario=crow['_resp'])
                    alertas_gerados.add(c_key)
                elif c_key not in alertas_gerados:
                    # C ou E: comentário sem NPS negativo na mesma aula
                    add(key,
                        f"Escreveu comentário em {fmt_data(c_data)}",
                        "Analisar comentário e dar retorno ao aluno",
                        crow['_topico_orig'], crow['_prof'],
                        comentario=crow['_resp'])
                    # Não adiciona ao alertas_gerados pois pode ter NPS negativo separado

    return alertas_por_aluno


def gerar_relatorio(df_canvas, alertas_nps, df_freq, desistentes_keys=None, turma_map=None):
    todos_keys = set(df_canvas['_key']) | set(alertas_nps.keys()) | set(df_freq['_key'])
    # Info base do canvas
    info_map = df_canvas.set_index('_key')[['Nome','Email','Curso','Turma']].to_dict('index')
    # Enriquecer com turma da frequência
    if turma_map:
        for k, turma in turma_map.items():
            if k in info_map:
                info_map[k]['Turma'] = turma
            else:
                info_map[k] = {'Nome': k.title(), 'Email': '', 'Curso': '', 'Turma': turma}
    for key in todos_keys:
        if key not in info_map:
            info_map[key] = {'Nome': key.title(), 'Email': '', 'Curso': '', 'Turma': ''}

    relatorio = []
    for key in sorted(todos_keys):
        if desistentes_keys and key in desistentes_keys:
            continue
        info = info_map[key]
        alertas_txt, acoes_txt, topicos_txt, profs_txt, comentarios_txt = [], [], [], [], []

        # Canvas
        c = df_canvas[df_canvas['_key'] == key]
        if not c.empty:
            dias = int(c.iloc[0]['Dias sem acesso'])
            alertas_txt.append(f"Sem acesso ao Canvas há {dias} dias")
            acoes_txt.append("Enviar link de acesso à plataforma")
            topicos_txt.append('')
            profs_txt.append('')
            comentarios_txt.append('')

        # NPS
        if key in alertas_nps:
            for al in alertas_nps[key]:
                alertas_txt.append(al['texto'])
                acoes_txt.append(al['acao'])
                topicos_txt.append(al.get('topico',''))
                profs_txt.append(al.get('professor',''))
                comentarios_txt.append(al.get('comentario',''))

        # Frequência
        f = df_freq[df_freq['_key'] == key]
        if not f.empty:
            alertas_txt.append(f"Ausente nas últimas 2 videoconferências ({f.iloc[0]['Ultimas_2_aulas']})")
            acoes_txt.append("Enviar data da próxima aula ao vivo")
            topicos_txt.append('')
            profs_txt.append('')
            comentarios_txt.append('')

        if alertas_txt:
            relatorio.append({
                'Curso':                 info['Curso'],
                'Turma':                 info['Turma'],
                'Nome':                  info['Nome'],
                'E-mail':                info['Email'],
                'Qtd. Alertas':          len(alertas_txt),
                'Alertas Identificados': ' | '.join(alertas_txt),
                'Ações Recomendadas':    ' | '.join(acoes_txt),
                'Tópico':                ' | '.join([t for t in topicos_txt if t]),
                'Professor':             ' | '.join([p for p in profs_txt if p]),
                'Comentário':            ' | '.join([c for c in comentarios_txt if c]),
            })

    df = pd.DataFrame(relatorio)
    return df.sort_values(['Curso','Turma','Qtd. Alertas','Nome'],
                          ascending=[True,True,False,True]).reset_index(drop=True)

def exportar_excel_bytes(df):
    import re as _re
    wb = Workbook()
    ws = wb.active; ws.title = "Relatório CS"
    VE="0F3D20"; VM="1B5E35"; BR="FFFFFF"; CR="F4F0E6"
    n_cols=10; lc=get_column_letter(n_cols)
    thin=Side(style='thin',color='E0DDD4')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    def hcell(cell,txt,bg,fg=BR,sz=10,bold=True,align='center'):
        cell.value=txt; cell.font=Font(bold=bold,size=sz,color=fg)
        cell.fill=PatternFill("solid",fgColor=bg)
        cell.alignment=Alignment(horizontal=align,vertical='center',wrap_text=True)
    ws.merge_cells(f'A1:{lc}1')
    hcell(ws['A1'],"RELATÓRIO DE ENGAJAMENTO — CUSTOMER SUCCESS REHAGRO",VE,sz=13)
    ws.row_dimensions[1].height=28
    ws.merge_cells(f'A2:{lc}2')
    hcell(ws['A2'],f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}   |   Total de alunos: {df['Nome'].nunique()}",VM,sz=9)
    ws.row_dimensions[2].height=18
    ws.append([])
    ws.merge_cells(f'A4:{lc}4')
    hcell(ws['A4'],"LEGENDA","EDF7EE",VE,sz=9,align='left')
    for i,txt in enumerate([
        "🟡 Canvas: Sem acesso há +20 dias  →  Enviar link",
        "🟠 Frequência: Ausente nas últimas 2 videoconferências  →  Enviar data da próxima aula",
        "🔴 NPS Detrator  →  Retomar feedback negativo",
        "🟢 Comentário registrado  →  Analisar e dar retorno ao aluno",
    ],5):
        ws.merge_cells(f'A{i}:{lc}{i}')
        ws[f'A{i}']=txt; ws[f'A{i}'].font=Font(size=8,color="5A5A4A")
        ws[f'A{i}'].fill=PatternFill("solid",fgColor=CR)
        ws[f'A{i}'].alignment=Alignment(horizontal='left',indent=2)
    ws.append([])
    headers=['Curso','Turma','Nome do Aluno','E-mail','Qtd. Alertas',
             'Alertas Identificados','Ações Recomendadas','Tópico','Professor','Comentário']
    ws.append(headers)
    hr=ws.max_row
    for ci,h in enumerate(headers,1):
        c=ws.cell(row=hr,column=ci)
        c.font=Font(bold=True,color=BR,size=9)
        c.fill=PatternFill("solid",fgColor=VE)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[hr].height=22

    def cor_alerta(alerta):
        a=str(alerta).lower()
        if 'canvas'      in a: return "FFF9C4"
        if 'videoconfer' in a: return "FFE0B2"
        if 'detrat'      in a: return "FFCDD2"
        if 'coment'      in a: return "E8F5E9"
        if 'presente'    in a: return "E3F2FD"
        return "F9F6EF"

    def expandir_linha(row):
        alertas    = [a.strip() for a in str(row.get('Alertas Identificados','')).split(' | ') if a.strip()]
        acoes      = [a.strip() for a in str(row.get('Ações Recomendadas','')).split(' | ') if a.strip()]
        topicos    = [t.strip() for t in str(row.get('Tópico','')).split(' | ')]
        profs      = [p.strip() for p in str(row.get('Professor','')).split(' | ')]
        comentarios= [c.strip() for c in str(row.get('Comentário','')).split(' | ')]

        # Separar alertas de plataforma (Canvas/Frequência) dos de NPS/Comentário
        plataforma, nps_coment = [], []
        for i, alerta in enumerate(alertas):
            acao = acoes[i] if i < len(acoes) else ''
            if any(x in alerta for x in ['Canvas', 'videoconfer']):
                plataforma.append((alerta, acao))
            else:
                nps_coment.append((i, alerta, acao))

        linhas = []

        # Canvas + Frequência → uma única linha combinada
        if plataforma:
            alertas_comb = ' | '.join(a for a, _ in plataforma)
            acoes_comb   = ' | '.join(ac for _, ac in plataforma)
            linhas.append({'alerta': alertas_comb, 'acao': acoes_comb,
                           'prof': '', 'topico': '', 'comentario': ''})

        # NPS / Comentários → cada um em linha própria
        tp_idx = 0
        for orig_i, alerta, acao in nps_coment:
            prof   = profs[tp_idx]       if tp_idx < len(profs)        else ''
            topico = topicos[tp_idx]     if tp_idx < len(topicos)      else ''
            coment = comentarios[tp_idx] if tp_idx < len(comentarios)  else ''
            tp_idx += 1
            linhas.append({'alerta': alerta, 'acao': acao,
                           'prof': prof, 'topico': topico, 'comentario': coment})

        return linhas if linhas else [{'alerta':'','acao':'','prof':'','topico':'','comentario':''}]

    for _,row in df.iterrows():
        linhas=expandir_linha(row)
        qtd=len(linhas)
        for li,linha in enumerate(linhas):
            ws.append([
                row['Curso']  if li==0 else '',
                row['Turma']  if li==0 else '',
                row['Nome']   if li==0 else '',
                row['E-mail'] if li==0 else '',
                qtd           if li==0 else '',
                linha['alerta'], linha['acao'], linha['prof'], linha['topico'], linha['comentario'],
            ])
            dr=ws.max_row; fc=cor_alerta(linha['alerta'])
            for ci in range(1,n_cols+1):
                c=ws.cell(row=dr,column=ci)
                if ci<=4:
                    c.fill=PatternFill("solid",fgColor="FFFFFF" if li==0 else "F8F8F8")
                    c.font=Font(size=8 if ci<=2 else 9,color=VM if (ci<=2 and li==0) else ("333333" if li==0 else "BBBBBB"),bold=(ci<=2 and li==0))
                elif ci==5:
                    c.fill=PatternFill("solid",fgColor="FFFFFF" if li==0 else "F8F8F8")
                    c.font=Font(bold=True,size=11 if li==0 else 9,color="333333" if li==0 else "BBBBBB")
                    c.alignment=Alignment(horizontal='center',vertical='center')
                elif ci in(6,7):
                    c.fill=PatternFill("solid",fgColor=fc); c.font=Font(size=9)
                elif ci==8:
                    c.fill=PatternFill("solid",fgColor="EEF7F0"); c.font=Font(size=8,color="2E7D32")
                elif ci==9:
                    c.fill=PatternFill("solid",fgColor="EEF7F0"); c.font=Font(size=8,color="1B5E35")
                elif ci==10:
                    tem=bool(str(linha['comentario']).strip())
                    c.fill=PatternFill("solid",fgColor="FFF8E1" if tem else "FFFFFF")
                    c.font=Font(size=8,color="5D4037" if tem else "AAAAAA",italic=not tem)
                if ci!=5: c.alignment=Alignment(vertical='center',wrap_text=True)
                c.border=border
            ws.row_dimensions[dr].height=30

    for ci,w in enumerate([14,16,28,26,8,55,38,28,36,45],1):
        ws.column_dimensions[get_column_letter(ci)].width=w

    ws2=wb.create_sheet("Resumo por Turma")
    for ci,w in enumerate([20,20,16,12,12,12],1): ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.append(["RESUMO POR TURMA","","","","",""])
    ws2['A1'].font=Font(bold=True,size=12,color=BR); ws2['A1'].fill=PatternFill("solid",fgColor=VE)
    ws2.merge_cells('A1:F1'); ws2['A1'].alignment=Alignment(horizontal='center')
    ws2.row_dimensions[1].height=24
    ws2.append(["Curso","Turma","Total Alunos","🟡 Canvas","🟠 Frequência","🔴🟢 NPS/Comentário"])
    hr2=ws2.max_row
    for ci in range(1,7):
        c=ws2.cell(row=hr2,column=ci); c.font=Font(bold=True,color=BR,size=9)
        c.fill=PatternFill("solid",fgColor=VM); c.alignment=Alignment(horizontal='center')
    ws2.row_dimensions[hr2].height=18
    for (curso,turma),g in df.groupby(['Curso','Turma']):
        n_c=g['Alertas Identificados'].str.contains('Canvas',na=False).sum()
        n_f=g['Alertas Identificados'].str.contains('videoconfer',na=False).sum()
        n_n=g['Alertas Identificados'].str.contains('etrat|oment|resente',na=False).sum()
        ws2.append([curso,turma,len(g),n_c,n_f,n_n])
        r=ws2.max_row
        for ci in range(1,7):
            c=ws2.cell(row=r,column=ci); c.font=Font(size=9); c.border=border
            c.alignment=Alignment(horizontal='center' if ci>2 else 'left')
        ws2.row_dimensions[r].height=16
    ws2.append(["TOTAL","",df['Nome'].nunique(),
                df['Alertas Identificados'].str.contains('Canvas',na=False).sum(),
                df['Alertas Identificados'].str.contains('videoconfer',na=False).sum(),
                df['Alertas Identificados'].str.contains('etrat|oment|resente',na=False).sum()])
    r=ws2.max_row
    for ci in range(1,7):
        c=ws2.cell(row=r,column=ci); c.font=Font(bold=True,size=9,color=BR)
        c.fill=PatternFill("solid",fgColor=VE); c.alignment=Alignment(horizontal='center' if ci>2 else 'left')
    ws2.row_dimensions[r].height=18
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def enviar_email(excel_bytes,destinatarios,data_hoje,total,criticos,atencao,monitorar):
    api_key=st.secrets.get("SENDGRID_API_KEY","")
    if not api_key: return False,"Chave SendGrid não configurada."
    lista=list(set(destinatarios+DESTINATARIOS_FIXOS))
    assunto=f"Relatório CS Rehagro — Engajamento {data_hoje}"
    html=f"""<div style="font-family:Outfit,Arial,sans-serif;max-width:600px;margin:0 auto;background:#F4F0E6;">
      <div style="background:#0F3D20;padding:36px 44px;border-bottom:3px solid #C8A951;">
        <p style="color:#C8A951;font-size:9px;font-weight:700;letter-spacing:4px;text-transform:uppercase;margin:0 0 12px">Customer Success · Rehagro</p>
        <h1 style="color:#fff;font-size:28px;margin:0;font-family:Georgia,serif;letter-spacing:-0.5px;">Relatório de Engajamento</h1>
        <p style="color:rgba(255,255,255,0.5);margin:8px 0 0;font-size:13px;">Gerado em {data_hoje}</p>
      </div>
      <div style="background:#fff;padding:36px 44px;border:1px solid #E0DDD4;">
        <p style="color:#444;font-size:14px;margin:0 0 28px;line-height:1.6;">Segue em anexo o relatório semanal de engajamento com alertas, ações recomendadas e comentários dos alunos.</p>
        <table width="100%" cellpadding="0" cellspacing="8">
          <tr>
            <td style="background:#EDF7EE;border-radius:8px;padding:18px 12px;text-align:center;border-bottom:3px solid #2E7D32">
              <div style="font-size:32px;font-weight:800;color:#0F3D20;font-family:Georgia,serif;">{total}</div>
              <div style="font-size:9px;color:#888;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px;">Total</div>
            </td>
            <td style="background:#FFEBEE;border-radius:8px;padding:18px 12px;text-align:center;border-bottom:3px solid #C62828">
              <div style="font-size:32px;font-weight:800;color:#C62828;font-family:Georgia,serif;">{criticos}</div>
              <div style="font-size:9px;color:#888;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px;">🔴 Críticos</div>
            </td>
            <td style="background:#FFF3E0;border-radius:8px;padding:18px 12px;text-align:center;border-bottom:3px solid #E65100">
              <div style="font-size:32px;font-weight:800;color:#E65100;font-family:Georgia,serif;">{atencao}</div>
              <div style="font-size:9px;color:#888;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px;">🟠 Atenção</div>
            </td>
            <td style="background:#FFFDE7;border-radius:8px;padding:18px 12px;text-align:center;border-bottom:3px solid #F57F17">
              <div style="font-size:32px;font-weight:800;color:#F57F17;font-family:Georgia,serif;">{monitorar}</div>
              <div style="font-size:9px;color:#888;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px;">🟡 Monitorar</div>
            </td>
          </tr>
        </table>
        <p style="color:#999;font-size:11px;margin-top:28px;border-top:1px solid #F0EDE4;padding-top:16px;">O arquivo Excel em anexo contém o relatório completo.</p>
      </div>
      <div style="background:#F4F0E6;padding:16px;text-align:center;border:1px solid #E0DDD4;border-top:none;">
        <p style="color:#aaa;font-size:10px;margin:0;letter-spacing:1px;">REHAGRO · CUSTOMER SUCCESS · AGENTE DE ENGAJAMENTO</p>
      </div>
    </div>"""
    encoded=base64.b64encode(excel_bytes).decode()
    nome_arq=f"relatorio_cs_{data_hoje.replace('/','')}.xlsx"
    message=Mail(from_email=REMETENTE,to_emails=lista[0],subject=assunto,html_content=html)
    if len(lista)>1:
        for dest in lista[1:]: message.add_cc(dest)
    message.attachment=Attachment(FileContent(encoded),FileName(nome_arq),
        FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),Disposition("attachment"))
    try:
        sg=SendGridAPIClient(api_key); res=sg.send(message)
        if res.status_code in [200,201,202]: return True,f"E-mail enviado para {len(lista)} destinatário(s)."
        return False,f"Erro SendGrid: status {res.status_code}"
    except Exception as e: return False,str(e)


# ── UI ────────────────────────────────────────────────────

# HERO
st.markdown("""
<div class="rh-hero">
  <div class="rh-hero-bg"></div>
  <div class="rh-hero-grid"></div>
  <div class="rh-hero-inner">
    <p class="rh-eyebrow">Rehagro · Customer Success</p>
    <h1 class="rh-hero-title">MONITORAMENTO <span>DE ALUNOS</span></h1>
    <p class="rh-hero-sub">Identifique alunos desengajados e saiba exatamente qual ação tomar — em segundos.</p>
    <div class="rh-hero-pills">
      <span class="rh-pill">Canvas AVA</span>
      <span class="rh-pill">NPS Avaliações</span>
      <span class="rh-pill">Frequência ao Vivo</span>
      <span class="rh-pill">Comentários</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="rh-body">', unsafe_allow_html=True)

col_esq, col_dir = st.columns([1, 1], gap="large")

with col_esq:
    st.markdown('<p class="rh-section">Como exportar do Power BI</p>', unsafe_allow_html=True)
    st.markdown("""
    <div>
      <div class="rh-dash-card" style="border-radius:12px 12px 0 0">
        <div class="rh-dash-num" style="font-size:1.1rem!important">Dashboard 01</div>
        <div class="rh-dash-title">Acesso ao Canvas (AVA)</div>
        <div class="rh-dash-desc">
          Relatório <b>Rehagro - Canvas</b> → página <b>Acesso ao Canvas-Ok</b><br><br>
          <span class="rh-tag">Status Usuário Curso = Ativo</span>
          <span class="rh-tag">Função na disciplina = Aluno</span>
          <span class="rh-tag">Curso = seus cursos</span>
          <span class="rh-tag">Turma = suas turmas</span><br><br>
          Exporte em formato <b>Dados Resumidos</b> → Excel
          <p class="rh-note">💡 Múltiplos cursos e turmas são suportados.</p>
        </div>
      </div>
      <div class="rh-dash-card">
        <div class="rh-dash-num" style="font-size:1.1rem!important">Dashboard 02</div>
        <div class="rh-dash-title">NPS Médio por Aluno</div>
        <div class="rh-dash-desc">
          Relatório <b>Rehagro Educação - Avaliação de Aula</b> → página <b>Avaliações de aula/aluno</b><br><br>
          <span class="rh-tag">Curso = seus cursos</span>
          <span class="rh-tag">Turma = suas turmas</span>
          <span class="rh-tag">Ano resposta = ano atual</span>
          <span class="rh-tag">Ano/mês resposta = período</span>
          <span class="rh-tag">Tipo de aula = On-line ao vivo</span><br><br>
          Exporte a tabela <b>NPS médio/aluno</b> → Excel
          <p class="rh-note">💡 Extraia da página "NPS médio/aluno" do dashboard.</p>
        </div>
      </div>
      <div class="rh-dash-card">
        <div class="rh-dash-num" style="font-size:1.1rem!important">Dashboard 03</div>
        <div class="rh-dash-title">Frequência nas Aulas ao Vivo</div>
        <div class="rh-dash-desc">
          Relatório <b>Rehagro Alunado</b> → página <b>Análise de Frequência e Faltas</b><br><br>
          <span class="rh-tag">Turma = suas turmas</span>
          <span class="rh-tag">Data/Aula = período desejado</span><br><br>
          Exporte a <b>Tabela de frequência</b> → Excel
        </div>
      </div>
      <div class="rh-dash-card" style="border-radius:0 0 12px 12px">
        <div class="rh-opt-badge">Opcional</div>
        <div class="rh-dash-num" style="font-size:1.1rem!important">Dashboard 04</div>
        <div class="rh-dash-title">Comentários das Aulas</div>
        <div class="rh-dash-desc">
          Relatório <b>Rehagro Educação - Avaliação de Aula</b> → página <b>Tabela Comentários</b><br><br>
          <span class="rh-tag">Área = sua área</span>
          <span class="rh-tag">Formato Curso = Online</span>
          <span class="rh-tag">Curso = seus cursos</span>
          <span class="rh-tag">Tipo de aula = On-line ao vivo</span>
          <span class="rh-tag">Ano_aula = ano atual</span><br><br>
          Exporte a <b>Tabela Comentários</b> → Excel
          <p class="rh-note">💡 Inclui data, tópico, professor e texto do comentário no relatório.</p>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

with col_dir:
    st.markdown('<p class="rh-section">Envie os arquivos</p>', unsafe_allow_html=True)

    f_canvas = st.file_uploader("1 — Canvas · Acesso ao AVA",           type=["xlsx"], key="canvas")
    f_nps    = st.file_uploader("2 — NPS · Avaliações de aula",          type=["xlsx"], key="nps")
    f_freq   = st.file_uploader("3 — Frequência · Aulas ao vivo",        type=["xlsx"], key="freq")
    f_coment = st.file_uploader("4 — Comentários · Opcional",            type=["xlsx"], key="coment")

    st.markdown('<p class="rh-section" style="margin-top:28px">Envio por e-mail</p>', unsafe_allow_html=True)
    email_usuario = st.text_input("Seu e-mail — receberá cópia junto à lista fixa:",
                                  placeholder="seuemail@rehagro.com.br")
    with st.expander("Ver destinatários fixos"):
        for d in DESTINATARIOS_FIXOS: st.write(f"• {d}")

    st.markdown('<div class="rh-divider"></div>', unsafe_allow_html=True)

    obrigatorios = f_canvas and f_nps and f_freq
    if not obrigatorios:
        faltando = [n for f,n in [(f_canvas,"Canvas"),(f_nps,"NPS"),(f_freq,"Frequência")] if not f]
        st.info(f"Aguardando: **{', '.join(faltando)}**")
    else:
        st.success("✅ Arquivos prontos — clique para gerar e enviar.")
        if st.button("Gerar e Enviar Relatório →", type="primary", use_container_width=True):
            with st.spinner("Analisando dados..."):
                try:
                    dc                              = carregar_canvas(f_canvas)
                    df_fr, desist_keys, turma_map, df_freq_ativo = carregar_frequencia(f_freq)
                    df_nps_raw                      = carregar_nps(f_nps, desistentes_keys=desist_keys)
                    df_co                           = carregar_comentarios(f_coment) if f_coment else None
                    alertas_nps                     = gerar_alertas_nps(df_nps_raw, df_co, df_freq_ativo)
                    df_rel                          = gerar_relatorio(dc, alertas_nps, df_fr, desist_keys, turma_map)

                    df_alunos = df_rel.drop_duplicates(subset=['Nome','Turma'])
                    criticos  = len(df_alunos[df_alunos['Qtd. Alertas']>=4])
                    atencao   = len(df_alunos[df_alunos['Qtd. Alertas'].between(2,3)])
                    monitorar = len(df_alunos[df_alunos['Qtd. Alertas']==1])
                    total_al  = df_rel['Nome'].nunique()

                    st.markdown(f"""
                    <div class="rh-metrics">
                      <div class="rh-metric m-total">
                        <div class="rh-metric-num" style="color:#0F3D20">{total_al}</div>
                        <div class="rh-metric-label">Total</div>
                      </div>
                      <div class="rh-metric m-crit">
                        <div class="rh-metric-num" style="color:#C62828">{criticos}</div>
                        <div class="rh-metric-label">🔴 Críticos</div>
                      </div>
                      <div class="rh-metric m-atenc">
                        <div class="rh-metric-num" style="color:#E65100">{atencao}</div>
                        <div class="rh-metric-label">🟠 Atenção</div>
                      </div>
                      <div class="rh-metric m-mon">
                        <div class="rh-metric-num" style="color:#F57F17">{monitorar}</div>
                        <div class="rh-metric-label">🟡 Monitorar</div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)

                    turmas = sorted(df_rel['Turma'].dropna().unique().tolist())
                    df_view = df_rel
                    if len(turmas) > 1:
                        sel = st.multiselect("Filtrar por turma:", turmas, default=turmas)
                        df_view = df_rel[df_rel['Turma'].isin(sel)]

                    st.markdown('<p class="rh-section">Alunos desengajados</p>', unsafe_allow_html=True)
                    cols_view = ['Curso','Turma','Nome','Qtd. Alertas','Alertas Identificados','Ações Recomendadas','Tópico','Professor']
                    st.dataframe(df_view[cols_view], use_container_width=True, hide_index=True, height=320)

                    excel_bytes = exportar_excel_bytes(df_rel)
                    data_hoje   = datetime.now().strftime('%d/%m/%Y')
                    dests = [email_usuario.strip()] if email_usuario and "@" in email_usuario else []
                    if SENDGRID_OK:
                        ok,msg = enviar_email(excel_bytes,dests,data_hoje,len(df_rel),criticos,atencao,monitorar)
                        if ok:
                            st.success(f"📧 {msg}")
                        else:
                            st.warning(f"⚠️ E-mail não enviado: {msg}")
                    else:
                        st.warning("⚠️ Biblioteca SendGrid não instalada.")

                    st.download_button("⬇  Baixar Relatório Excel", data=excel_bytes,
                        file_name=f"relatorio_cs_{datetime.now().strftime('%d%m%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                except Exception as e:
                    st.error(f"Erro ao processar: {e}")
                    st.info("Verifique se os arquivos corretos foram enviados com os filtros indicados.")

st.markdown('</div>', unsafe_allow_html=True)
st.markdown("""
<div class="rh-footer">
  <span>Rehagro</span>
  <span class="rh-footer-dot"></span>
  <span>Customer Success</span>
  <span class="rh-footer-dot"></span>
  <span>Agente de Engajamento</span>
  <span class="rh-footer-dot"></span>
  <span>© 2026</span>
</div>
""", unsafe_allow_html=True)
