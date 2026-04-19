import streamlit as st
import pandas as pd
import re
import unicodedata
import io
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

try:
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import (Mail, Attachment, FileContent,
                                        FileName, FileType, Disposition, Cc)
    SENDGRID_OK = True
except ImportError:
    SENDGRID_OK = False

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800;900&family=Inter:wght@400;500;600&display=swap');

:root {
  --g:     #1B3D2A;
  --g2:    #2a5c3f;
  --gold:  #C8A532;
  --cream: #F2EDE4;
  --c2:    #fdfcf9;
  --c3:    #f8f6f0;
  --txt:   #1c1c1c;
  --txt2:  #4a4a4a;
  --muted: #888;
  --bd:    #ddd8ce;
  --r:     14px;
}

*, *::before, *::after { box-sizing: border-box; }
html, body, [class*="css"] {
    font-family: 'Inter', sans-serif !important;
    background: var(--cream) !important;
    color: var(--txt) !important;
}
.stApp { background: var(--cream) !important; }
#MainMenu, footer, header { visibility: hidden; }
[data-testid="stAppViewContainer"] { padding-top: 0 !important; }
[data-testid="block-container"]    { padding-top: 0 !important; }
button[aria-label="Close sidebar"],
button[aria-label="Fechar barra lateral"] { display: none !important; }

/* ── HERO ── */
.rh-hero {
    background: #1B3D2A;
    padding: 18px 40px 24px; margin: -1rem -1rem 0 -1rem;
    position: relative; overflow: hidden;
}
.rh-hero-diag {
    position: absolute; inset: 0; opacity: .05;
    background-image: repeating-linear-gradient(45deg,#fff 0,#fff 1px,transparent 0,transparent 50%);
    background-size: 12px 12px; pointer-events: none;
}
.rh-hero-inner { position: relative; z-index: 1; }
.rh-hero-eyebrow {
    color: #C8A532; font-family: 'Montserrat', sans-serif;
    font-weight: 700; font-size: 11px; letter-spacing: 2px; text-transform: uppercase;
    margin: 0 0 8px; display: flex; align-items: center; gap: 8px;
}
.rh-hero-eyebrow::before { content: ''; display: inline-block; width: 24px; height: 2px; background: #C8A532; }
.rh-hero-h1 {
    font-family: 'Montserrat', sans-serif !important; font-weight: 900 !important;
    font-size: 34px !important; color: #C8A532 !important;
    letter-spacing: 1px !important; margin: 0 0 6px !important; text-transform: uppercase !important;
}
.rh-hero-sub { color: rgba(255,255,255,.65); font-size: 13px; max-width: 520px; line-height: 1.5; margin: 0; }
.rh-hero-pills { display: flex; gap: 8px; margin-top: 14px; flex-wrap: wrap; }
.rh-hero-pill {
    padding: 4px 12px; border: 1.5px solid rgba(255,255,255,.3); border-radius: 20px;
    color: rgba(255,255,255,.8); font-size: 11px; font-weight: 600;
    letter-spacing: .5px; text-transform: uppercase;
}

/* ── Section header ── */
.rh-section {
    font-family: 'Montserrat', sans-serif; font-size: 11px; font-weight: 800;
    letter-spacing: 2px; text-transform: uppercase; color: #555;
    margin: 0 0 16px; display: flex; align-items: center; gap: 10px;
}
.rh-section::after { content: ''; flex: 1; height: 1px; background: linear-gradient(to right, #ddd, transparent); }

/* ── Dashboard cards — novo design ── */
.rh-dash-card {
    background: #fff;
    border-radius: 0;
    padding: 0;
    margin-bottom: 2px;
    box-shadow: 0 1px 4px rgba(0,0,0,.04);
    overflow: hidden;
}
.rh-dash-card:first-of-type { border-radius: var(--r) var(--r) 0 0; }
.rh-dash-card:last-of-type  { border-radius: 0 0 var(--r) var(--r); margin-bottom: 0; }

.rh-dash-card-bar {
    height: 3px;
    width: 100%;
}
.rh-dash-card-bar.c-blue   { background: #2563eb; }
.rh-dash-card-bar.c-purple { background: #7c3aed; }
.rh-dash-card-bar.c-green  { background: #16a34a; }

.rh-dash-card-body {
    padding: 18px 20px 16px;
    display: flex;
    gap: 14px;
    align-items: flex-start;
}

.rh-dash-num-circle {
    width: 32px;
    height: 32px;
    border-radius: 50%;
    flex-shrink: 0;
    display: flex;
    align-items: center;
    justify-content: center;
    font-family: 'Montserrat', sans-serif;
    font-weight: 800;
    font-size: 13px;
    color: #fff;
    margin-top: 2px;
}
.rh-dash-num-circle.c-blue   { background: #2563eb; }
.rh-dash-num-circle.c-purple { background: #7c3aed; }
.rh-dash-num-circle.c-green  { background: #16a34a; }

.rh-dash-content { flex: 1; }

.rh-dash-eyebrow {
    font-family: 'Montserrat', sans-serif;
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    margin-bottom: 4px;
}
.rh-dash-eyebrow.c-blue   { color: #2563eb; }
.rh-dash-eyebrow.c-purple { color: #7c3aed; }
.rh-dash-eyebrow.c-green  { color: #16a34a; }

.rh-dash-title {
    font-family: 'Montserrat', sans-serif;
    font-size: 15px;
    font-weight: 700;
    color: var(--txt);
    margin-bottom: 10px;
    line-height: 1.3;
}

.rh-dash-source {
    font-size: 12px;
    color: var(--muted);
    margin-bottom: 10px;
    display: flex;
    align-items: center;
    gap: 6px;
}
.rh-dash-source::before {
    content: '';
    display: inline-block;
    width: 12px;
    height: 12px;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' fill='none' viewBox='0 0 24 24' stroke='%23888'%3E%3Cpath stroke-linecap='round' stroke-linejoin='round' stroke-width='2' d='M9 19v-6a2 2 0 00-2-2H5a2 2 0 00-2 2v6a2 2 0 002 2h2a2 2 0 002-2zm0 0V9a2 2 0 012-2h2a2 2 0 012 2v10m-6 0a2 2 0 002 2h2a2 2 0 002-2m0 0V5a2 2 0 012-2h2a2 2 0 012 2v14a2 2 0 01-2 2h-2a2 2 0 01-2-2z'/%3E%3C/svg%3E");
    background-size: contain;
    background-repeat: no-repeat;
    flex-shrink: 0;
}

.rh-filters-label {
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 6px;
    display: flex;
    align-items: center;
    gap: 6px;
}
.rh-filters-label::before {
    content: '▼';
    font-size: 8px;
}

.rh-tag-filled {
    display: inline-block;
    background: #1B3D2A;
    color: #fff;
    font-size: 11px;
    font-weight: 600;
    padding: 3px 10px;
    border-radius: 20px;
    margin: 2px 3px 2px 0;
}
.rh-tag-outline {
    display: inline-block;
    background: transparent;
    color: #444;
    border: 1.5px solid #ccc;
    font-size: 11px;
    font-weight: 500;
    padding: 3px 10px;
    border-radius: 20px;
    margin: 2px 3px 2px 0;
}

.rh-export-line {
    display: flex;
    align-items: center;
    gap: 6px;
    font-size: 12px;
    color: var(--txt2);
    margin-top: 10px;
}
.rh-export-line a, .rh-export-line strong {
    color: var(--g);
    font-weight: 600;
    text-decoration: none;
}

.rh-note-tip {
    background: rgba(200,165,50,.08);
    border-left: 3px solid var(--gold);
    border-radius: 0 6px 6px 0;
    padding: 6px 12px;
    margin-top: 10px;
    font-size: 12px;
    color: #7a6010;
    font-style: italic;
}
.rh-note-tip::before { content: '⚡ '; }

.rh-opt-badge {
    display: inline-block;
    background: rgba(200,165,50,.12);
    border: 1px solid rgba(200,165,50,.4);
    color: var(--gold);
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    padding: 2px 10px;
    border-radius: 4px;
    margin-bottom: 4px;
}

/* ── Upload sidebar — novo design ── */
.rh-upload-panel {
    background: #fff;
    border-radius: var(--r);
    overflow: hidden;
    box-shadow: 0 1px 8px rgba(0,0,0,.06);
    margin-bottom: 16px;
}
.rh-upload-header {
    padding: 14px 18px 12px;
    border-bottom: 1px solid var(--bd);
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.rh-upload-header-title {
    font-family: 'Montserrat', sans-serif;
    font-size: 11px;
    font-weight: 800;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: #555;
    display: flex;
    align-items: center;
    gap: 8px;
}
.rh-upload-progress {
    font-family: 'Montserrat', sans-serif;
    font-size: 12px;
    font-weight: 700;
    color: var(--gold);
}

.rh-upload-row {
    display: flex;
    align-items: center;
    padding: 12px 18px;
    border-bottom: 1px solid #f0ede6;
    gap: 12px;
}
.rh-upload-row:last-child { border-bottom: none; }
.rh-upload-row-num {
    width: 22px;
    height: 22px;
    border-radius: 50%;
    background: #f0ede6;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 11px;
    font-weight: 700;
    color: #888;
    flex-shrink: 0;
}
.rh-upload-row-num.done { background: var(--g); color: #fff; }
.rh-upload-row-name {
    flex: 1;
    font-size: 13px;
    font-weight: 600;
    color: var(--txt);
}
.rh-upload-row-sub {
    font-size: 11px;
    color: var(--muted);
    font-weight: 400;
}
.rh-upload-row-opt {
    font-size: 10px;
    color: var(--gold);
    font-weight: 700;
    letter-spacing: 1px;
    text-transform: uppercase;
    background: rgba(200,165,50,.1);
    border: 1px solid rgba(200,165,50,.3);
    padding: 1px 7px;
    border-radius: 4px;
}

/* ── Status aguardando pills ── */
.rh-waiting {
    display: flex;
    align-items: center;
    gap: 8px;
    flex-wrap: wrap;
    padding: 10px 0;
    font-size: 12px;
    color: var(--muted);
}
.rh-waiting-dot {
    width: 8px; height: 8px;
    border-radius: 50%;
    background: var(--gold);
    flex-shrink: 0;
}
.rh-waiting-pill {
    background: rgba(200,165,50,.1);
    border: 1px solid rgba(200,165,50,.3);
    color: #7a6010;
    font-size: 11px;
    font-weight: 600;
    padding: 2px 10px;
    border-radius: 20px;
}
.rh-waiting-pill.done {
    background: rgba(22,163,74,.08);
    border-color: rgba(22,163,74,.25);
    color: #16a34a;
}

/* ── E-mail section ── */
.rh-email-section {
    background: #fff;
    border-radius: var(--r);
    overflow: hidden;
    box-shadow: 0 1px 8px rgba(0,0,0,.06);
    margin-bottom: 16px;
}
.rh-email-header {
    padding: 14px 18px 12px;
    border-bottom: 1px solid var(--bd);
    font-family: 'Montserrat', sans-serif;
    font-size: 11px;
    font-weight: 800;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: #555;
    display: flex;
    align-items: center;
    gap: 8px;
}
.rh-email-body { padding: 16px 18px; }
.rh-email-label {
    font-size: 12px;
    color: var(--muted);
    margin-bottom: 8px;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] label,
[data-testid="stFileUploader"] > div > label {
    font-size: 13px !important; font-weight: 600 !important;
    color: var(--g) !important; margin-bottom: 4px !important;
}
[data-testid="stFileUploader"] {
    background: #fff !important; border-radius: 10px !important;
    border: 1.5px dashed rgba(27,61,42,.2) !important;
    padding: 6px 10px !important; margin-bottom: 12px !important;
}
[data-testid="stFileUploader"]:focus-within {
    border-color: var(--gold) !important;
    box-shadow: 0 0 0 3px rgba(200,165,50,.12) !important;
}

/* ── Text input ── */
[data-testid="stTextInput"] label { font-size: 12px !important; font-weight: 600 !important; color: #555 !important; }
[data-testid="stTextInput"] input {
    border-radius: 8px !important; border: 1.5px solid #ddd !important;
    font-size: 13px !important; font-family: 'Inter', sans-serif !important;
}

/* ── Buttons ── */
.stButton > button {
    font-family: 'Montserrat', sans-serif !important; font-weight: 700 !important;
    border-radius: 8px !important; font-size: 13px !important;
}
.stButton > button[kind="primary"] {
    background: var(--g) !important; color: #fff !important;
    border: none !important; padding: 12px 24px !important;
    box-shadow: 0 4px 14px rgba(27,61,42,.3) !important;
}
.stButton > button[kind="secondary"] {
    background: #fff !important; border: 1.5px solid var(--bd) !important; color: #555 !important;
}
.stDownloadButton > button {
    font-family: 'Montserrat', sans-serif !important; font-weight: 700 !important;
    border-radius: 9px !important; background: var(--g) !important;
    color: #fff !important; border: none !important;
    font-size: 13px !important; box-shadow: 0 4px 14px rgba(27,61,42,.3) !important;
}

/* ── Metrics ── */
.rh-metrics { display: grid; grid-template-columns: repeat(4,1fr); gap: 12px; margin: 20px 0; }
.rh-metric {
    background: #fff; border-radius: var(--r); padding: 18px 14px;
    text-align: center; border: 1px solid var(--bd); position: relative; overflow: hidden;
}
.rh-metric::after { content: ''; position: absolute; bottom: 0; left: 0; right: 0; height: 3px; }
.rh-metric.m-total::after { background: var(--g); }
.rh-metric.m-crit::after  { background: #dc2626; }
.rh-metric.m-atenc::after { background: #d97706; }
.rh-metric.m-mon::after   { background: #ca8a04; }
.rh-metric-num {
    font-family: 'Montserrat', sans-serif; font-weight: 900;
    font-size: 2.8rem; line-height: 1; margin-bottom: 4px;
}
.rh-metric-label { font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 1.5px; color: var(--muted); }

/* ── Misc ── */
.rh-body { padding: 24px 0 0; }
.rh-divider { height: 1px; background: linear-gradient(90deg, var(--bd), transparent); margin: 20px 0; opacity: .5; }
[data-testid="stExpander"] { border: 1px solid var(--bd) !important; border-radius: 8px !important; background: #fff !important; }
[data-testid="stDataFrame"] { border-radius: var(--r) !important; overflow: hidden !important; border: 1px solid var(--bd) !important; }
[data-testid="stAlert"] { border-radius: 8px !important; font-size: 13px !important; }
.rh-footer {
    display: flex; align-items: center; justify-content: center; gap: 16px;
    padding: 32px 0 16px; border-top: 1px solid var(--bd); margin-top: 40px;
    color: var(--muted); font-size: 12px;
}
.rh-footer-dot { width: 4px; height: 4px; border-radius: 50%; background: var(--gold); display: inline-block; }

/* ── Seção de resultados ── */
.rh-results-header {
    display: flex;
    align-items: flex-end;
    justify-content: space-between;
    margin: 32px 0 20px;
    padding-bottom: 12px;
    border-bottom: 2px solid var(--bd);
}
.rh-results-eyebrow {
    font-family: 'Montserrat', sans-serif;
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: var(--gold);
    margin-bottom: 4px;
}
.rh-results-title {
    font-family: 'Montserrat', sans-serif;
    font-size: 28px;
    font-weight: 900;
    color: var(--g);
    margin: 0;
    line-height: 1.1;
}

/* ── Cards de métricas — novo design ── */
.rh-metrics-new {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 12px;
    margin-bottom: 24px;
}
.rh-metric-new {
    background: #fff;
    border-radius: var(--r);
    padding: 20px 18px;
    border: 1px solid var(--bd);
    position: relative;
    overflow: hidden;
}
.rh-metric-new-top {
    display: flex;
    align-items: flex-start;
    justify-content: space-between;
    margin-bottom: 12px;
}
.rh-metric-new-label {
    font-family: 'Montserrat', sans-serif;
    font-size: 10px;
    font-weight: 800;
    letter-spacing: 1.5px;
    text-transform: uppercase;
}
.rh-metric-new-icon {
    width: 32px;
    height: 32px;
    border-radius: 8px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 16px;
}
.rh-metric-new-num {
    font-family: 'Montserrat', sans-serif;
    font-weight: 900;
    font-size: 48px;
    line-height: 1;
    margin-bottom: 4px;
}
.rh-metric-new-sub {
    font-size: 11px;
    color: var(--muted);
    font-weight: 500;
}
.rh-metric-new.m-total { background: #f0fdf4; border-color: rgba(22,163,74,.15); }
.rh-metric-new.m-total .rh-metric-new-label { color: #166534; }
.rh-metric-new.m-total .rh-metric-new-icon  { background: rgba(22,163,74,.12); }
.rh-metric-new.m-total .rh-metric-new-num   { color: #14532d; }
.rh-metric-new.m-crit { background: #fff5f5; border-color: rgba(220,38,38,.15); }
.rh-metric-new.m-crit .rh-metric-new-label { color: #991b1b; }
.rh-metric-new.m-crit .rh-metric-new-icon  { background: rgba(220,38,38,.10); }
.rh-metric-new.m-crit .rh-metric-new-num   { color: #dc2626; }
.rh-metric-new.m-atenc { background: #fffbeb; border-color: rgba(217,119,6,.15); }
.rh-metric-new.m-atenc .rh-metric-new-label { color: #92400e; }
.rh-metric-new.m-atenc .rh-metric-new-icon  { background: rgba(217,119,6,.10); }
.rh-metric-new.m-atenc .rh-metric-new-num   { color: #d97706; }
.rh-metric-new.m-mon { background: #fefce8; border-color: rgba(202,138,4,.15); }
.rh-metric-new.m-mon .rh-metric-new-label { color: #854d0e; }
.rh-metric-new.m-mon .rh-metric-new-icon  { background: rgba(202,138,4,.10); }
.rh-metric-new.m-mon .rh-metric-new-num   { color: #ca8a04; }

/* ── Filtro de turma ── */
.rh-turma-filter {
    background: #fff;
    border-radius: var(--r);
    border: 1px solid var(--bd);
    padding: 14px 18px;
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 16px;
    flex-wrap: wrap;
}
.rh-turma-filter-label {
    font-size: 13px;
    font-weight: 600;
    color: var(--txt2);
    white-space: nowrap;
}

/* ── Tabela de alunos ── */
.rh-table-wrap {
    background: #fff;
    border-radius: var(--r);
    border: 1px solid var(--bd);
    overflow: hidden;
    margin-bottom: 20px;
}
.rh-table-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 14px 20px;
    border-bottom: 1px solid var(--bd);
    background: #fafaf8;
}
.rh-table-header-title {
    font-family: 'Montserrat', sans-serif;
    font-size: 11px;
    font-weight: 800;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: var(--txt);
    display: flex;
    align-items: center;
    gap: 8px;
}
.rh-table-header-count {
    font-size: 12px;
    color: var(--muted);
    font-weight: 500;
}
.rh-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
}
.rh-table thead th {
    padding: 10px 16px;
    text-align: left;
    font-family: 'Montserrat', sans-serif;
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: var(--muted);
    border-bottom: 1px solid var(--bd);
    background: #fafaf8;
    white-space: nowrap;
}
.rh-table tbody tr {
    border-bottom: 1px solid #f0ede6;
    transition: background .1s;
}
.rh-table tbody tr:last-child { border-bottom: none; }
.rh-table tbody tr:hover { background: #fafaf8; }
.rh-table tbody td {
    padding: 14px 16px;
    vertical-align: middle;
    color: var(--txt);
}
.rh-table-curso  { font-size: 12px; color: var(--muted); }
.rh-table-turma  { font-size: 12px; color: var(--muted); }
.rh-avatar {
    width: 32px;
    height: 32px;
    border-radius: 50%;
    background: #e8e4dc;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    font-size: 11px;
    font-weight: 700;
    color: #555;
    margin-right: 10px;
    flex-shrink: 0;
    vertical-align: middle;
}
.rh-table-nome-cell {
    display: flex;
    align-items: center;
}
.rh-table-nome-text {
    font-weight: 600;
    color: var(--txt);
}
.rh-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 700;
    margin-left: 8px;
}
.rh-badge.critico  { background: rgba(220,38,38,.08);  color: #dc2626; border: 1px solid rgba(220,38,38,.2);  }
.rh-badge.atencao  { background: rgba(217,119,6,.08);  color: #d97706; border: 1px solid rgba(217,119,6,.2);  }
.rh-badge.monitorar{ background: rgba(202,138,4,.08);  color: #ca8a04; border: 1px solid rgba(202,138,4,.2);  }
.rh-table-alertas-cell {
    display: flex;
    align-items: center;
    white-space: nowrap;
}
.rh-alertas-num {
    font-family: 'Montserrat', sans-serif;
    font-weight: 800;
    font-size: 15px;
    min-width: 20px;
}
.rh-alertas-num.critico  { color: #dc2626; }
.rh-alertas-num.atencao  { color: #d97706; }
.rh-alertas-num.monitorar{ color: #ca8a04; }
.rh-table-comentario { font-size: 12px; color: var(--muted); }

/* ── Botões finais ── */
.rh-action-buttons {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 12px;
    margin-top: 8px;
    margin-bottom: 8px;
}
</style>
""", unsafe_allow_html=True)

# ── Configurações de e-mail ──────────────────────────────
DESTINATARIOS_FIXOS = ["rafael.ferraz@rehagro.edu.br"]
REMETENTE = "rafael.ferraz@rehagro.edu.br"

# ── Helpers ──────────────────────────────────────────────
def _norm(n):
    n = str(n).strip().lower()
    n = unicodedata.normalize('NFD', n)
    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
    n = re.sub(r'\s+', ' ', n)
    return n

_MESES = {'janeiro':1,'fevereiro':2,'março':3,'abril':4,'maio':5,'junho':6,
          'julho':7,'agosto':8,'setembro':9,'outubro':10,'novembro':11,'dezembro':12}

def _parse_freq_date(s):
    m = re.match(r'(\d{2}/\d{2}/\d{4})', str(s))
    return pd.to_datetime(m.group(1), dayfirst=True) if m else pd.NaT

def _montar_data_nps(row):
    try:
        ano = int(row['Data da aula - Ano'])
        mes = _MESES.get(str(row['Data da aula - Mês']).strip().lower())
        dia = int(row['Data da aula - Dia'])
        if mes: return pd.Timestamp(year=ano, month=mes, day=dia)
    except: pass
    return pd.NaT


def carregar_canvas(arquivo):
    df = pd.read_excel(arquivo, skiprows=2)
    df.columns = df.columns.str.strip()
    col_nome   = next((c for c in df.columns if 'NOME' in c.upper()), None)
    col_email  = next((c for c in df.columns if 'E-MAIL' in c.upper() or 'EMAIL' in c.upper()), None)
    col_dias   = next((c for c in df.columns if 'DIAS' in c.upper()), None)
    col_curso  = next((c for c in df.columns if c.upper() == 'CURSO'), None)
    col_turma  = next((c for c in df.columns if c.upper() == 'TURMA'), None)
    col_funcao = next((c for c in df.columns if 'FUN' in c.upper() and 'DISCIPLINA' in c.upper()), None)
    if not col_nome or not col_dias:
        raise ValueError("Arquivo Canvas: colunas não encontradas.")
    if col_funcao:
        df = df[df[col_funcao].astype(str).str.upper().str.contains('ALUNO', na=False)].copy()
    df['_key']  = df[col_nome].apply(_norm)
    df['_dias'] = pd.to_numeric(df[col_dias], errors='coerce')
    cols = {'_key':'_key', col_nome:'Nome', '_dias':'Dias sem acesso'}
    if col_email: cols[col_email] = 'Email'
    if col_curso: cols[col_curso] = 'Curso'
    if col_turma: cols[col_turma] = 'Turma'
    res = df[df['_dias'] > 20][list(cols.keys())].copy().rename(columns=cols)
    if 'Email' not in res.columns: res['Email'] = ''
    if 'Curso' not in res.columns: res['Curso'] = ''
    if 'Turma' not in res.columns: res['Turma'] = ''
    return res[['_key','Nome','Email','Curso','Turma','Dias sem acesso']]


def carregar_frequencia(arquivo):
    df = pd.read_excel(arquivo, skiprows=2)
    df.columns = df.columns.str.strip()
    col_aluno  = next((c for c in df.columns if c.upper() == 'ALUNO'), None)
    col_status = next((c for c in df.columns if 'STATUS' in c.upper() or 'PRESEN' in c.upper()), None)
    col_data   = next((c for c in df.columns if 'DATA' in c.upper() or 'PARTE' in c.upper()), None)
    col_turma  = next((c for c in df.columns if c.upper() == 'TURMA'), None)
    if not col_aluno or not col_status:
        raise ValueError("Arquivo Frequência: colunas não encontradas.")
    df['_key'] = df[col_aluno].apply(_norm)
    df['_data'] = df[col_data].apply(_parse_freq_date) if col_data else pd.NaT
    desistentes_keys = set(
        df[df[col_status].astype(str).str.upper() == 'DESISTENTE']['_key'].unique()
    )
    turma_map = {}
    if col_turma:
        for _, row in df.iterrows():
            k = row['_key']
            if k not in turma_map and pd.notna(row[col_turma]):
                turma_map[k] = str(row[col_turma]).strip()
    df_a = df[~df['_key'].isin(desistentes_keys) & (df[col_status].astype(str).str.strip() != '-')].copy()
    df_a = df_a.sort_values(['_key', '_data'])
    ausentes = []
    for nome, g in df_a.groupby('_key'):
        if len(g) < 2: continue
        last2 = g.tail(2)
        status_last2 = last2[col_status].astype(str).str.upper().tolist()
        if status_last2 == ['AUSENTE', 'AUSENTE']:
            datas = last2[col_data].tolist() if col_data else ['N/D','N/D']
            ausentes.append({'_key': nome, 'Ultimas_2_aulas': f"{datas[0]} e {datas[1]}"})
    df_ausentes = pd.DataFrame(ausentes) if ausentes else pd.DataFrame(columns=['_key','Ultimas_2_aulas'])
    return df_ausentes, desistentes_keys, turma_map, df_a


def carregar_comentarios(arquivo):
    df = pd.read_excel(arquivo, skiprows=2)
    df.columns = df.columns.str.strip()
    col_aluno  = next((c for c in df.columns if c.upper() == 'ALUNO'), None)
    col_data   = next((c for c in df.columns if 'DATA DA AULA' in c.upper() or c.upper() == 'DATA DA AULA'), None)
    col_topico = next((c for c in df.columns if 'TÓPICO' in c.upper() or 'TOPICO' in c.upper()), None)
    col_prof   = next((c for c in df.columns if 'PROFESSOR' in c.upper()), None)
    col_resp   = next((c for c in df.columns if 'RESPOSTA' in c.upper()), None)
    if not col_aluno or not col_resp:
        raise ValueError("Arquivo Comentários: colunas não encontradas.")
    df = df[df[col_resp].notna() & (df[col_resp].astype(str).str.strip() != '')].copy()
    df['_key']    = df[col_aluno].apply(_norm)
    df['_data']   = pd.to_datetime(df[col_data], errors='coerce').dt.normalize() if col_data else pd.NaT
    df['_topico'] = df[col_topico].astype(str).str.strip().str.lower() if col_topico else ''
    df['_prof']   = df[col_prof].astype(str).str.strip() if col_prof else ''
    df['_topico_orig'] = df[col_topico].astype(str).str.strip() if col_topico else ''
    df['_resp']   = df[col_resp].astype(str).str.strip()
    return df[['_key','_data','_topico','_topico_orig','_prof','_resp']]


def carregar_nps(arquivo, desistentes_keys=None):
    df = pd.read_excel(arquivo, skiprows=2)
    df.columns = df.columns.str.strip()
    col_aluno  = next((c for c in df.columns if c.upper() == 'ALUNO'), None)
    col_nps    = next((c for c in df.columns if 'NPS' in c.upper() and 'REA' in c.upper()), None)
    col_topico = next((c for c in df.columns if 'TÓPICO' in c.upper() or 'TOPICO' in c.upper()), None)
    col_prof   = next((c for c in df.columns if 'PROFESSOR' in c.upper()), None)
    if not col_aluno or not col_nps:
        raise ValueError("Arquivo NPS: colunas não encontradas.")
    df['_data'] = df.apply(_montar_data_nps, axis=1)
    df['_key']  = df[col_aluno].apply(_norm)
    df['_topico']      = df[col_topico].astype(str).str.strip().str.lower() if col_topico else ''
    df['_topico_orig'] = df[col_topico].astype(str).str.strip() if col_topico else ''
    df['_prof']        = df[col_prof].astype(str).str.strip() if col_prof else ''
    df['_nps']         = pd.to_numeric(df[col_nps], errors='coerce')
    if desistentes_keys:
        df = df[~df['_key'].isin(desistentes_keys)].copy()
    df = df.sort_values(['_key','_data'])
    return df[['_key','_data','_topico','_topico_orig','_prof','_nps']]


def gerar_alertas_nps(df_nps, df_coment, df_freq_ativo):
    alertas_por_aluno = {}

    def fmt_data(d):
        try: return pd.Timestamp(d).strftime('%d/%m/%Y')
        except: return str(d)

    def add(key, texto, acao, topico='', professor='', comentario=''):
        alertas_por_aluno.setdefault(key, []).append(
            {'texto': texto, 'acao': acao, 'topico': topico, 'professor': professor, 'comentario': comentario}
        )

    coment_map = {}
    if df_coment is not None and not df_coment.empty:
        for _, row in df_coment.iterrows():
            k = (row['_key'], row['_data'], row['_topico'])
            coment_map[k] = {'prof': row['_prof'], 'resp': row['_resp'],
                             'topico_orig': row['_topico_orig']}

    freq_por_aluno = {}
    if df_freq_ativo is not None and not df_freq_ativo.empty:
        for nome, g in df_freq_ativo.groupby('_key'):
            freq_por_aluno[nome] = g.sort_values('_data')

    todos_keys = set(df_nps['_key'].unique()) | set(df_coment['_key'].unique() if df_coment is not None else [])

    for key in todos_keys:
        nps_aluno    = df_nps[df_nps['_key'] == key].sort_values('_data')
        coment_aluno = df_coment[df_coment['_key'] == key] if df_coment is not None else pd.DataFrame()
        freq_aluno   = freq_por_aluno.get(key, pd.DataFrame())
        alertas_gerados = set()

        respondidas = nps_aluno[nps_aluno['_nps'].notna()]
        ult2_resp   = respondidas.tail(2)
        if len(ult2_resp) == 2 and all(ult2_resp['_nps'] < 0):
            detalhes = []
            for _, r in ult2_resp.iterrows():
                detalhes.append(f"{fmt_data(r['_data'])} · {r['_topico_orig']} · Prof. {r['_prof']}")
                alertas_gerados.add((r['_data'], r['_topico']))
            topicos  = ' · '.join([r['_topico_orig'] for _,r in ult2_resp.iterrows()])
            profs    = ' · '.join([r['_prof'] for _,r in ult2_resp.iterrows()])
            add(key, f"Detrator nos últimos 2 encontros: {' || '.join(detalhes)}",
                "Retomar feedback negativo da avaliação", topicos, profs)

        if not freq_aluno.empty and len(freq_aluno) >= 2:
            ult2_freq = freq_aluno.tail(2)
            status_ult2 = ult2_freq['Primeiro StatusPresenca'].astype(str).str.upper().tolist() \
                if 'Primeiro StatusPresenca' in ult2_freq.columns else []
            if status_ult2 == ['PRESENTE', 'PRESENTE']:
                datas_freq = ult2_freq['_data'].tolist()
                nps_nessas = nps_aluno[nps_aluno['_data'].isin(datas_freq) & nps_aluno['_nps'].notna()]
                if nps_nessas.empty:
                    datas_str = ' e '.join([fmt_data(d) for d in datas_freq])
                    add(key, f"Presente nas últimas 2 aulas sem avaliar ({datas_str})",
                        "Incentivar participação nas avaliações de aula")

        if not freq_aluno.empty and len(freq_aluno) >= 2:
            ult   = freq_aluno.iloc[-1]
            penul = freq_aluno.iloc[-2]
            status_cols = [c for c in freq_aluno.columns if 'STATUS' in c.upper() or 'PRESEN' in c.upper()]
            if status_cols:
                sc = status_cols[0]
                if str(ult[sc]).upper() == 'AUSENTE':
                    data_penul  = penul['_data']
                    nps_penul   = nps_aluno[nps_aluno['_data'] == data_penul]
                    det_penul   = nps_penul[nps_penul['_nps'] < 0]
                    if not det_penul.empty:
                        r = det_penul.iloc[0]
                        chave_aula = (r['_data'], r['_topico'])
                        if chave_aula not in alertas_gerados:
                            add(key,
                                f"Detrator na penúltima aula ({fmt_data(r['_data'])} · {r['_topico_orig']} · Prof. {r['_prof']}) e ausente na última ({fmt_data(ult['_data'])})",
                                "Retomar feedback e verificar ausência", r['_topico_orig'], r['_prof'])
                            alertas_gerados.add(chave_aula)

        if not coment_aluno.empty:
            for _, crow in coment_aluno.iterrows():
                c_data   = crow['_data']
                c_topico = crow['_topico']
                c_key    = (c_data, c_topico)
                nps_mesma = nps_aluno[(nps_aluno['_data'] == c_data) &
                                      (nps_aluno['_topico'] == c_topico) &
                                      (nps_aluno['_nps'] < 0)]
                if not nps_mesma.empty and c_key not in alertas_gerados:
                    r = nps_mesma.iloc[0]
                    add(key, f"Escreveu comentário em {fmt_data(c_data)} (com avaliação negativa)",
                        "Retomar feedback negativo e comentário da avaliação",
                        crow['_topico_orig'], crow['_prof'], comentario=crow['_resp'])
                    alertas_gerados.add(c_key)
                elif c_key not in alertas_gerados:
                    add(key, f"Escreveu comentário em {fmt_data(c_data)}",
                        "Analisar comentário e dar retorno ao aluno",
                        crow['_topico_orig'], crow['_prof'], comentario=crow['_resp'])

    return alertas_por_aluno


def gerar_relatorio(df_canvas, alertas_nps, df_freq, desistentes_keys=None, turma_map=None):
    todos_keys = set(df_canvas['_key']) | set(alertas_nps.keys()) | set(df_freq['_key'])
    info_map = df_canvas.set_index('_key')[['Nome','Email','Curso','Turma']].to_dict('index')
    if turma_map:
        for k, turma in turma_map.items():
            if k in info_map:
                info_map[k]['Turma'] = turma
            else:
                info_map[k] = {'Nome': k.title(), 'Email': '', 'Curso': '', 'Turma': turma}
    for key in todos_keys:
        if key not in info_map:
            info_map[key] = {'Nome': key.title(), 'Email': '', 'Curso': '', 'Turma': ''}

    relatorio = []
    for key in sorted(todos_keys):
        if desistentes_keys and key in desistentes_keys:
            continue
        info = info_map[key]
        alertas_txt, acoes_txt, topicos_txt, profs_txt, comentarios_txt = [], [], [], [], []

        c = df_canvas[df_canvas['_key'] == key]
        if not c.empty:
            dias = int(c.iloc[0]['Dias sem acesso'])
            alertas_txt.append(f"Sem acesso ao Canvas há {dias} dias")
            acoes_txt.append("Enviar link de acesso à plataforma")
            topicos_txt.append(''); profs_txt.append(''); comentarios_txt.append('')

        if key in alertas_nps:
            for al in alertas_nps[key]:
                alertas_txt.append(al['texto']); acoes_txt.append(al['acao'])
                topicos_txt.append(al.get('topico','')); profs_txt.append(al.get('professor',''))
                comentarios_txt.append(al.get('comentario',''))

        f = df_freq[df_freq['_key'] == key]
        if not f.empty:
            alertas_txt.append(f"Ausente nas últimas 2 videoconferências ({f.iloc[0]['Ultimas_2_aulas']})")
            acoes_txt.append("Enviar data da próxima aula ao vivo")
            topicos_txt.append(''); profs_txt.append(''); comentarios_txt.append('')

        if alertas_txt:
            relatorio.append({
                'Curso': info['Curso'], 'Turma': info['Turma'],
                'Nome': info['Nome'], 'E-mail': info['Email'],
                'Qtd. Alertas': len(alertas_txt),
                'Alertas Identificados': ' | '.join(alertas_txt),
                'Ações Recomendadas':    ' | '.join(acoes_txt),
                'Tópico':    ' | '.join([t for t in topicos_txt if t]),
                'Professor': ' | '.join([p for p in profs_txt if p]),
                'Comentário':' | '.join([c for c in comentarios_txt if c]),
            })

    df = pd.DataFrame(relatorio)
    return df.sort_values(['Curso','Turma','Qtd. Alertas','Nome'],
                          ascending=[True,True,False,True]).reset_index(drop=True)

def exportar_excel_bytes(df):
    wb = Workbook()
    ws = wb.active; ws.title = "Relatório CS"
    VE="0F3D20"; VM="1B5E35"; BR="FFFFFF"; CR="F4F0E6"
    n_cols=10; lc=get_column_letter(n_cols)
    thin=Side(style='thin',color='E0DDD4')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    def hcell(cell,txt,bg,fg=BR,sz=10,bold=True,align='center'):
        cell.value=txt; cell.font=Font(bold=bold,size=sz,color=fg)
        cell.fill=PatternFill("solid",fgColor=bg)
        cell.alignment=Alignment(horizontal=align,vertical='center',wrap_text=True)
    ws.merge_cells(f'A1:{lc}1')
    hcell(ws['A1'],"RELATÓRIO DE ENGAJAMENTO — CUSTOMER SUCCESS REHAGRO",VE,sz=13)
    ws.row_dimensions[1].height=52
    try:
        logo_xl = XLImage("Logo-Rehagro-chapada-branca.png")
        logo_xl.height = 61; logo_xl.width = 210
        logo_xl.anchor = get_column_letter(n_cols - 1) + "1"
        ws.add_image(logo_xl)
    except Exception:
        pass
    ws.merge_cells(f'A2:{lc}2')
    hcell(ws['A2'],f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}   |   Total de alunos: {df['Nome'].nunique()}",VM,sz=9)
    ws.row_dimensions[2].height=18
    ws.append([])
    ws.merge_cells(f'A4:{lc}4')
    hcell(ws['A4'],"LEGENDA","EDF7EE",VE,sz=9,align='left')
    for i,txt in enumerate(["🟡 Canvas: Sem acesso há +20 dias  →  Enviar link",
        "🟠 Frequência: Ausente nas últimas 2 videoconferências  →  Enviar data da próxima aula",
        "🔴 NPS Detrator  →  Retomar feedback negativo",
        "🟢 Comentário registrado  →  Analisar e dar retorno ao aluno"],5):
        ws.merge_cells(f'A{i}:{lc}{i}')
        ws[f'A{i}']=txt; ws[f'A{i}'].font=Font(size=11,color="5A5A4A")
        ws[f'A{i}'].fill=PatternFill("solid",fgColor=CR)
        ws[f'A{i}'].alignment=Alignment(horizontal='left',indent=2)
    ws.append([])
    headers=['Curso','Turma','Nome do Aluno','E-mail','Qtd. Alertas',
             'Alertas Identificados','Ações Recomendadas','Tópico','Professor','Comentário']
    ws.append(headers)
    hr=ws.max_row
    for ci,h in enumerate(headers,1):
        c=ws.cell(row=hr,column=ci)
        c.font=Font(bold=True,color=BR,size=9)
        c.fill=PatternFill("solid",fgColor=VE)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[hr].height=22

    def cor_alerta(alerta):
        a=str(alerta).lower()
        if 'canvas' in a: return "FFF9C4"
        if 'videoconfer' in a: return "FFE0B2"
        if 'detrat' in a: return "FFCDD2"
        if 'coment' in a: return "E8F5E9"
        if 'presente' in a: return "E3F2FD"
        return "F9F6EF"

    def expandir_linha(row):
        alertas    = [a.strip() for a in str(row.get('Alertas Identificados','')).split(' | ') if a.strip()]
        acoes      = [a.strip() for a in str(row.get('Ações Recomendadas','')).split(' | ') if a.strip()]
        topicos    = [t.strip() for t in str(row.get('Tópico','')).split(' | ')]
        profs      = [p.strip() for p in str(row.get('Professor','')).split(' | ')]
        comentarios= [c.strip() for c in str(row.get('Comentário','')).split(' | ')]
        plataforma, nps_coment = [], []
        for i, alerta in enumerate(alertas):
            acao = acoes[i] if i < len(acoes) else ''
            if any(x in alerta for x in ['Canvas', 'videoconfer']):
                plataforma.append((alerta, acao))
            else:
                nps_coment.append((i, alerta, acao))
        linhas = []
        if plataforma:
            linhas.append({'alerta': ' | '.join(a for a,_ in plataforma),
                           'acao':   ' | '.join(ac for _,ac in plataforma),
                           'prof': '', 'topico': '', 'comentario': ''})
        tp_idx = 0
        for orig_i, alerta, acao in nps_coment:
            linhas.append({'alerta': alerta, 'acao': acao,
                           'prof':      profs[tp_idx]       if tp_idx < len(profs)        else '',
                           'topico':    topicos[tp_idx]     if tp_idx < len(topicos)      else '',
                           'comentario':comentarios[tp_idx] if tp_idx < len(comentarios)  else ''})
            tp_idx += 1
        return linhas if linhas else [{'alerta':'','acao':'','prof':'','topico':'','comentario':''}]

    for _,row in df.iterrows():
        linhas=expandir_linha(row); qtd=row['Qtd. Alertas']
        for li,linha in enumerate(linhas):
            ws.append([row['Curso'] if li==0 else '', row['Turma'] if li==0 else '',
                       row['Nome']  if li==0 else '', row['E-mail'] if li==0 else '',
                       qtd if li==0 else '',
                       linha['alerta'],linha['acao'],linha['topico'],linha['prof'],linha['comentario']])
            dr=ws.max_row; fc=cor_alerta(linha['alerta'])
            for ci in range(1,n_cols+1):
                c=ws.cell(row=dr,column=ci)
                if ci<=4:
                    c.fill=PatternFill("solid",fgColor="FFFFFF" if li==0 else "F8F8F8")
                    c.font=Font(size=8 if ci<=2 else 9,color=VM if (ci<=2 and li==0) else ("333333" if li==0 else "BBBBBB"),bold=(ci<=2 and li==0))
                elif ci==5:
                    c.fill=PatternFill("solid",fgColor="FFFFFF" if li==0 else "F8F8F8")
                    c.font=Font(bold=True,size=11 if li==0 else 9,color="333333" if li==0 else "BBBBBB")
                    c.alignment=Alignment(horizontal='center',vertical='center')
                elif ci in(6,7):
                    c.fill=PatternFill("solid",fgColor=fc); c.font=Font(size=9)
                elif ci==8:
                    c.fill=PatternFill("solid",fgColor="EEF7F0"); c.font=Font(size=8,color="2E7D32")
                elif ci==9:
                    c.fill=PatternFill("solid",fgColor="EEF7F0"); c.font=Font(size=8,color="1B5E35")
                elif ci==10:
                    tem=bool(str(linha['comentario']).strip())
                    c.fill=PatternFill("solid",fgColor="FFF8E1" if tem else "FFFFFF")
                    c.font=Font(size=8,color="5D4037" if tem else "AAAAAA",italic=not tem)
                if ci!=5: c.alignment=Alignment(vertical='center',wrap_text=True)
                c.border=border
            ws.row_dimensions[dr].height=30

    for ci,w in enumerate([14,16,28,26,8,55,38,28,36,45],1):
        ws.column_dimensions[get_column_letter(ci)].width=w

    ws2=wb.create_sheet("Resumo por Turma")
    for ci,w in enumerate([20,20,16,12,12,12],1): ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.append(["RESUMO POR TURMA","","","","",""])
    ws2['A1'].font=Font(bold=True,size=12,color=BR); ws2['A1'].fill=PatternFill("solid",fgColor=VE)
    ws2.merge_cells('A1:F1'); ws2['A1'].alignment=Alignment(horizontal='center')
    ws2.row_dimensions[1].height=24
    ws2.append(["Curso","Turma","Total Alunos","🟡 Canvas","🟠 Frequência","🔴🟢 NPS/Comentário"])
    hr2=ws2.max_row
    for ci in range(1,7):
        c=ws2.cell(row=hr2,column=ci); c.font=Font(bold=True,color=BR,size=9)
        c.fill=PatternFill("solid",fgColor=VM); c.alignment=Alignment(horizontal='center')
    ws2.row_dimensions[hr2].height=18
    for (curso,turma),g in df.groupby(['Curso','Turma']):
        n_c=g['Alertas Identificados'].str.contains('Canvas',na=False).sum()
        n_f=g['Alertas Identificados'].str.contains('videoconfer',na=False).sum()
        n_n=g['Alertas Identificados'].str.contains('etrat|oment|resente',na=False).sum()
        ws2.append([curso,turma,len(g),n_c,n_f,n_n])
        r=ws2.max_row
        for ci in range(1,7):
            c=ws2.cell(row=r,column=ci); c.font=Font(size=9); c.border=border
            c.alignment=Alignment(horizontal='center' if ci>2 else 'left')
        ws2.row_dimensions[r].height=16
    ws2.append(["TOTAL","",df['Nome'].nunique(),
                df['Alertas Identificados'].str.contains('Canvas',na=False).sum(),
                df['Alertas Identificados'].str.contains('videoconfer',na=False).sum(),
                df['Alertas Identificados'].str.contains('etrat|oment|resente',na=False).sum()])
    r=ws2.max_row
    for ci in range(1,7):
        c=ws2.cell(row=r,column=ci); c.font=Font(bold=True,size=9,color=BR)
        c.fill=PatternFill("solid",fgColor=VE); c.alignment=Alignment(horizontal='center' if ci>2 else 'left')
    ws2.row_dimensions[r].height=18
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def enviar_email(excel_bytes,destinatarios,data_hoje,total,criticos,atencao,monitorar):
    api_key=st.secrets.get("SENDGRID_API_KEY","")
    if not api_key: return False,"Chave SendGrid não configurada."
    lista=list(set(destinatarios+DESTINATARIOS_FIXOS))
    assunto=f"Relatório CS Rehagro — Engajamento {data_hoje}"
    html=f"""<div style="font-family:Outfit,Arial,sans-serif;max-width:600px;margin:0 auto;background:#F4F0E6;">
      <div style="background:#0F3D20;padding:36px 44px;border-bottom:3px solid #C8A951;">
        <p style="color:#C8A951;font-size:9px;font-weight:700;letter-spacing:4px;text-transform:uppercase;margin:0 0 12px">Customer Success · Rehagro</p>
        <h1 style="color:#fff;font-size:28px;margin:0;font-family:Georgia,serif;letter-spacing:-0.5px;">Relatório de Engajamento</h1>
        <p style="color:rgba(255,255,255,0.5);margin:8px 0 0;font-size:13px;">Gerado em {data_hoje}</p>
      </div>
      <div style="background:#fff;padding:36px 44px;border:1px solid #E0DDD4;">
        <p style="color:#444;font-size:14px;margin:0 0 28px;line-height:1.6;">Segue em anexo o relatório semanal de engajamento com alertas, ações recomendadas e comentários dos alunos.</p>
        <table width="100%" cellpadding="0" cellspacing="8"><tr>
          <td style="background:#EDF7EE;border-radius:8px;padding:18px 12px;text-align:center;border-bottom:3px solid #2E7D32"><div style="font-size:32px;font-weight:800;color:#0F3D20;font-family:Georgia,serif;">{total}</div><div style="font-size:9px;color:#888;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px;">Total</div></td>
          <td style="background:#FFEBEE;border-radius:8px;padding:18px 12px;text-align:center;border-bottom:3px solid #C62828"><div style="font-size:32px;font-weight:800;color:#C62828;font-family:Georgia,serif;">{criticos}</div><div style="font-size:9px;color:#888;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px;">🔴 Críticos</div></td>
          <td style="background:#FFF3E0;border-radius:8px;padding:18px 12px;text-align:center;border-bottom:3px solid #E65100"><div style="font-size:32px;font-weight:800;color:#E65100;font-family:Georgia,serif;">{atencao}</div><div style="font-size:9px;color:#888;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px;">🟠 Atenção</div></td>
          <td style="background:#FFFDE7;border-radius:8px;padding:18px 12px;text-align:center;border-bottom:3px solid #F57F17"><div style="font-size:32px;font-weight:800;color:#F57F17;font-family:Georgia,serif;">{monitorar}</div><div style="font-size:9px;color:#888;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px;">🟡 Monitorar</div></td>
        </tr></table>
        <p style="color:#999;font-size:11px;margin-top:28px;border-top:1px solid #F0EDE4;padding-top:16px;">O arquivo Excel em anexo contém o relatório completo.</p>
      </div>
      <div style="background:#F4F0E6;padding:16px;text-align:center;border:1px solid #E0DDD4;border-top:none;">
        <p style="color:#aaa;font-size:10px;margin:0;letter-spacing:1px;">REHAGRO · CUSTOMER SUCCESS · AGENTE DE ENGAJAMENTO</p>
      </div></div>"""
    encoded=base64.b64encode(excel_bytes).decode()
    nome_arq=f"relatorio_cs_{data_hoje.replace('/','')}.xlsx"
    message=Mail(from_email=REMETENTE,to_emails=lista[0],subject=assunto,html_content=html)
    if len(lista)>1:
        for dest in lista[1:]: message.add_cc(dest)
    message.attachment=Attachment(FileContent(encoded),FileName(nome_arq),
        FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),Disposition("attachment"))
    try:
        sg=SendGridAPIClient(api_key); res=sg.send(message)
        if res.status_code in [200,201,202]: return True,f"E-mail enviado para {len(lista)} destinatário(s)."
        return False,f"Erro SendGrid: status {res.status_code}"
    except Exception as e: return False,str(e)


# ── UI ────────────────────────────────────────────────────

def _logo_b64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return ""

_logo_hero = _logo_b64("Logo-Rehagro-branca-transp.png")
_logo_img  = f'<img src="data:image/png;base64,{_logo_hero}" style="height:76px;opacity:0.92;" />' if _logo_hero else ""

st.markdown(f"""
<div class="rh-hero">
  <div class="rh-hero-diag"></div>
  <div class="rh-hero-inner" style="display:flex;justify-content:space-between;align-items:center;">
    <div>
      <p class="rh-hero-eyebrow">Rehagro · Customer Success</p>
      <h1 class="rh-hero-h1" style="color:#C8A532 !important">MONITORAMENTO DE ALUNOS</h1>
      <p class="rh-hero-sub">Identifique alunos desengajados e saiba exatamente qual ação tomar — em segundos.</p>
      <div class="rh-hero-pills">
        <span class="rh-hero-pill">Canvas AVA</span>
        <span class="rh-hero-pill">NPS Avaliações</span>
        <span class="rh-hero-pill">Frequência ao Vivo</span>
        <span class="rh-hero-pill">Comentários</span>
      </div>
    </div>
    <div style="flex-shrink:0;padding-left:24px;">{_logo_img}</div>
  </div>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="rh-body">', unsafe_allow_html=True)

if st.button("🔍  Comportamento do Aluno →", key="nav_comp"):
    st.switch_page("pages/comportamento_aluno.py")

col_esq, col_dir = st.columns([1, 1], gap="large")

with col_esq:
    st.markdown('<p class="rh-section">Como exportar do Power BI</p>', unsafe_allow_html=True)
    st.markdown("""
<div style="background:rgba(200,165,50,.08);border:1px solid rgba(200,165,50,.3);border-left:4px solid var(--gold);border-radius:8px;padding:14px 18px;margin-bottom:20px;">
  <div style="font-size:13px;font-weight:700;color:var(--g);margin-bottom:5px;">⚠️ Atenção antes de exportar</div>
  <div style="font-size:13px;color:#555;line-height:1.6;">Todos os arquivos devem ser exportados no formato <strong>Dados Resumidos</strong> no Power BI. Cada dashboard abaixo indica os <strong>filtros obrigatórios</strong> para a exportação correta.</div>
</div>

<div>
  <!-- DASHBOARD 01 -->
  <div class="rh-dash-card" style="border-radius:12px 12px 0 0">
    <div class="rh-dash-card-bar c-blue"></div>
    <div class="rh-dash-card-body">
      <div class="rh-dash-num-circle c-blue">01</div>
      <div class="rh-dash-content">
        <div class="rh-dash-eyebrow c-blue">Dashboard 01</div>
        <div class="rh-dash-title">Acesso ao Canvas (AVA)</div>
        <div class="rh-dash-source">Rehagro - Canvas → página <em>Acesso ao Canvas-Ok</em></div>
        <div class="rh-filters-label">Filtros obrigatórios</div>
        <div>
          <span class="rh-tag-filled">Status Usuário Curso</span>
          <span class="rh-tag-outline">Ativo</span>
          <span class="rh-tag-filled">Função na disciplina</span>
          <span class="rh-tag-outline">Aluno</span>
          <span class="rh-tag-filled">Curso</span>
          <span class="rh-tag-outline">seus cursos</span>
          <span class="rh-tag-filled">Turma</span>
          <span class="rh-tag-outline">suas turmas</span>
        </div>
        <div class="rh-export-line">✈ Exporte em formato <strong>Dados Resumidos</strong> → Excel</div>
        <div class="rh-note-tip">Múltiplos cursos e turmas são suportados.</div>
      </div>
    </div>
  </div>

  <!-- DASHBOARD 02 -->
  <div class="rh-dash-card">
    <div class="rh-dash-card-bar c-purple"></div>
    <div class="rh-dash-card-body">
      <div class="rh-dash-num-circle c-purple">02</div>
      <div class="rh-dash-content">
        <div class="rh-dash-eyebrow c-purple">Dashboard 02</div>
        <div class="rh-dash-title">NPS Médio por Aluno</div>
        <div class="rh-dash-source">Rehagro Educação - Avaliação de Aula → página <em>Avaliações de aula/aluno</em></div>
        <div class="rh-filters-label">Filtros obrigatórios</div>
        <div>
          <span class="rh-tag-filled">Curso</span>
          <span class="rh-tag-outline">seus cursos</span>
          <span class="rh-tag-filled">Turma</span>
          <span class="rh-tag-outline">suas turmas</span>
          <span class="rh-tag-filled">Ano resposta</span>
          <span class="rh-tag-outline">ano atual</span>
          <span class="rh-tag-filled">Ano/mês resposta</span>
          <span class="rh-tag-outline">período</span>
          <span class="rh-tag-filled">Tipo de aula</span>
          <span class="rh-tag-outline">On-line ao vivo</span>
        </div>
        <div class="rh-export-line">✈ Exporte a <strong>tabela NPS médio/aluno</strong> → Excel</div>
        <div class="rh-note-tip">Extraia da página "NPS médio/aluno" do dashboard.</div>
      </div>
    </div>
  </div>

  <!-- DASHBOARD 03 -->
  <div class="rh-dash-card">
    <div class="rh-dash-card-bar c-green"></div>
    <div class="rh-dash-card-body">
      <div class="rh-dash-num-circle c-green">03</div>
      <div class="rh-dash-content">
        <div class="rh-dash-eyebrow c-green">Dashboard 03</div>
        <div class="rh-dash-title">Frequência nas Aulas ao Vivo</div>
        <div class="rh-dash-source">Rehagro Alunado → página <em>Análise de Frequência e Faltas</em></div>
        <div class="rh-filters-label">Filtros obrigatórios</div>
        <div>
          <span class="rh-tag-filled">Turma</span>
          <span class="rh-tag-outline">suas turmas</span>
          <span class="rh-tag-filled">Data/Aula</span>
          <span class="rh-tag-outline">período desejado</span>
        </div>
        <div class="rh-export-line">✈ Exporte a <strong>Tabela de frequência</strong> → Excel</div>
      </div>
    </div>
  </div>

  <!-- DASHBOARD 04 — OPCIONAL -->
  <div class="rh-dash-card" style="border-radius:0 0 12px 12px">
    <div class="rh-dash-card-bar c-green"></div>
    <div class="rh-dash-card-body">
      <div class="rh-dash-num-circle c-green">04</div>
      <div class="rh-dash-content">
        <div class="rh-opt-badge">Opcional</div>
        <div class="rh-dash-eyebrow c-green">Dashboard 04</div>
        <div class="rh-dash-title">Comentários das Aulas</div>
        <div class="rh-dash-source">Rehagro Educação - Avaliação de Aula → página <em>Tabela Comentários</em></div>
        <div class="rh-filters-label">Filtros obrigatórios</div>
        <div>
          <span class="rh-tag-filled">Área</span>
          <span class="rh-tag-outline">sua área</span>
          <span class="rh-tag-filled">Formato Curso</span>
          <span class="rh-tag-outline">Online</span>
          <span class="rh-tag-filled">Curso</span>
          <span class="rh-tag-outline">seus cursos</span>
          <span class="rh-tag-filled">Tipo de aula</span>
          <span class="rh-tag-outline">On-line ao vivo</span>
          <span class="rh-tag-filled">Ano_aula</span>
          <span class="rh-tag-outline">ano atual</span>
        </div>
        <div class="rh-export-line">✈ Exporte a <strong>Tabela Comentários</strong> → Excel</div>
        <div class="rh-note-tip">Inclui data, tópico, professor e texto do comentário no relatório.</div>
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

with col_dir:
    st.markdown("""
<div class="rh-upload-panel">
  <div class="rh-upload-header">
    <div class="rh-upload-header-title">🔼 Envie os arquivos</div>
    <div class="rh-upload-progress">Progresso do envio <span id="prog">0/4</span></div>
  </div>
  <div class="rh-upload-row">
    <div class="rh-upload-row-num">1</div>
    <div style="flex:1"><div class="rh-upload-row-name">Canvas · Acesso ao AVA</div></div>
  </div>
  <div class="rh-upload-row">
    <div class="rh-upload-row-num">2</div>
    <div style="flex:1"><div class="rh-upload-row-name">NPS · Avaliações de aula</div></div>
  </div>
  <div class="rh-upload-row">
    <div class="rh-upload-row-num">3</div>
    <div style="flex:1"><div class="rh-upload-row-name">Frequência · Aulas ao vivo</div></div>
  </div>
  <div class="rh-upload-row" style="border-bottom:none">
    <div class="rh-upload-row-num">4</div>
    <div style="flex:1"><div class="rh-upload-row-name">Comentários <span class="rh-upload-row-sub">· Opcional</span></div></div>
    <span class="rh-upload-row-opt">Opcional</span>
  </div>
</div>
""", unsafe_allow_html=True)

    f_canvas = st.file_uploader("1 — Canvas · Acesso ao AVA",           type=["xlsx"], key="canvas")
    f_nps    = st.file_uploader("2 — NPS · Avaliações de aula",          type=["xlsx"], key="nps")
    f_freq   = st.file_uploader("3 — Frequência · Aulas ao vivo",        type=["xlsx"], key="freq")
    f_coment = st.file_uploader("4 — Comentários · Opcional",            type=["xlsx"], key="coment")

    st.markdown("""
<div class="rh-email-header" style="margin-top:24px;background:#fff;border-radius:12px 12px 0 0;border:1px solid var(--bd);padding:14px 18px;">
  ✉️ <span style="font-family:'Montserrat',sans-serif;font-size:11px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:#555;">Envio por E-mail</span>
</div>
""", unsafe_allow_html=True)
    email_usuario = st.text_input("Seu e-mail — receberá cópia junto à lista fixa:",
                                  placeholder="seuemail@rehagro.com.br")
    with st.expander("Ver destinatários fixos"):
        for d in DESTINATARIOS_FIXOS: st.write(f"• {d}")

    st.markdown('<div class="rh-divider"></div>', unsafe_allow_html=True)

    obrigatorios = f_canvas and f_nps and f_freq
    if not obrigatorios:
        faltando = [n for f,n in [(f_canvas,"Canvas"),(f_nps,"NPS"),(f_freq,"Frequência")] if not f]
        prontos  = [n for f,n in [(f_canvas,"Canvas"),(f_nps,"NPS"),(f_freq,"Frequência")] if f]
        pills_faltando = "".join(f'<span class="rh-waiting-pill">{n}</span>' for n in faltando)
        pills_prontos  = "".join(f'<span class="rh-waiting-pill done">✓ {n}</span>' for n in prontos)
        st.markdown(f"""
    <div class="rh-waiting">
      <div class="rh-waiting-dot"></div>
      <span style="font-weight:600;color:#555;">Aguardando:</span>
      {pills_faltando}
      {pills_prontos}
      <span style="margin-left:auto;font-size:11px;color:var(--muted);">{len(faltando)}/3 obrigatórios</span>
    </div>
    """, unsafe_allow_html=True)
    else:
        st.success("✅ Arquivos prontos — clique para gerar e enviar.")
        if st.button("Gerar e Enviar Relatório →", type="primary", use_container_width=True):
            with st.spinner("Analisando dados..."):
                try:
                    dc                              = carregar_canvas(f_canvas)
                    df_fr, desist_keys, turma_map, df_freq_ativo = carregar_frequencia(f_freq)
                    df_nps_raw                      = carregar_nps(f_nps, desistentes_keys=desist_keys)
                    df_co                           = carregar_comentarios(f_coment) if f_coment else None
                    alertas_nps                     = gerar_alertas_nps(df_nps_raw, df_co, df_freq_ativo)
                    df_rel                          = gerar_relatorio(dc, alertas_nps, df_fr, desist_keys, turma_map)

                    df_alunos = df_rel.drop_duplicates(subset=['Nome','Turma'])
                    criticos  = len(df_alunos[df_alunos['Qtd. Alertas']>=4])
                    atencao   = len(df_alunos[df_alunos['Qtd. Alertas'].between(2,3)])
                    monitorar = len(df_alunos[df_alunos['Qtd. Alertas']==1])
                    total_al  = df_rel['Nome'].nunique()

                    # ── Header da seção de resultados ──
                    st.markdown(f"""
<div class="rh-results-header">
  <div>
    <div class="rh-results-eyebrow">Relatório Gerado</div>
    <div class="rh-results-title">Alunos Desengajados</div>
  </div>
</div>
""", unsafe_allow_html=True)

                    # ── Cards de métricas — novo design ──
                    st.markdown(f"""
<div class="rh-metrics-new">
  <div class="rh-metric-new m-total">
    <div class="rh-metric-new-top">
      <div class="rh-metric-new-label">Total</div>
      <div class="rh-metric-new-icon">👥</div>
    </div>
    <div class="rh-metric-new-num">{total_al}</div>
    <div class="rh-metric-new-sub">alunos</div>
  </div>
  <div class="rh-metric-new m-crit">
    <div class="rh-metric-new-top">
      <div class="rh-metric-new-label">Críticos</div>
      <div class="rh-metric-new-icon">⚠️</div>
    </div>
    <div class="rh-metric-new-num">{criticos}</div>
    <div class="rh-metric-new-sub">alunos</div>
  </div>
  <div class="rh-metric-new m-atenc">
    <div class="rh-metric-new-top">
      <div class="rh-metric-new-label">Atenção</div>
      <div class="rh-metric-new-icon">🕐</div>
    </div>
    <div class="rh-metric-new-num">{atencao}</div>
    <div class="rh-metric-new-sub">alunos</div>
  </div>
  <div class="rh-metric-new m-mon">
    <div class="rh-metric-new-top">
      <div class="rh-metric-new-label">Monitorar</div>
      <div class="rh-metric-new-icon">⭐</div>
    </div>
    <div class="rh-metric-new-num">{monitorar}</div>
    <div class="rh-metric-new-sub">alunos</div>
  </div>
</div>
""", unsafe_allow_html=True)

                    # ── Filtro de turma ──
                    turmas = sorted(df_rel['Turma'].dropna().unique().tolist())
                    df_view = df_rel
                    if len(turmas) > 1:
                        st.markdown("""
<div style="background:#fff;border-radius:12px;border:1px solid var(--bd);
            padding:12px 18px;margin-bottom:8px;font-size:13px;
            font-weight:600;color:var(--txt2);">
  Filtrar por turma:
</div>
""", unsafe_allow_html=True)
                        sel = st.multiselect("", turmas, default=turmas, label_visibility="collapsed")
                        df_view = df_rel[df_rel['Turma'].isin(sel)]
                        total_view = df_view['Nome'].nunique()
                    else:
                        total_view = total_al

                    # ── Funções auxiliares para a tabela ──
                    def _iniciais(nome):
                        partes = str(nome).strip().split()
                        if len(partes) >= 2:
                            return (partes[0][0] + partes[1][0]).upper()
                        elif partes:
                            return partes[0][:2].upper()
                        return "??"

                    def _severidade(qtd):
                        if qtd >= 4: return "critico", "Crítico"
                        if qtd >= 2: return "atencao", "Atenção"
                        return "monitorar", "Monitorar"

                    def _ultimo_comentario(alertas_str):
                        matches = re.findall(r'comentário em (\d{2}/\d{2}/\d{4})', str(alertas_str))
                        if matches:
                            return f"Escreveu comentário em {matches[-1]}"
                        return "—"

                    # ── Montar linhas da tabela ──
                    df_tabela = df_view.drop_duplicates(subset=['Nome','Turma']).sort_values(
                        ['Qtd. Alertas','Nome'], ascending=[False, True]
                    )
                    linhas_html = ""
                    for _, row in df_tabela.iterrows():
                        iniciais = _iniciais(row['Nome'])
                        sev_cls, sev_lbl = _severidade(row['Qtd. Alertas'])
                        qtd = int(row['Qtd. Alertas'])
                        comentario = _ultimo_comentario(row.get('Alertas Identificados', ''))
                        curso = str(row.get('Curso', '')).strip() or '—'
                        turma = str(row.get('Turma', '')).strip() or '—'
                        linhas_html += f"""
    <tr>
      <td class="rh-table-curso">{curso}</td>
      <td class="rh-table-turma">{turma}</td>
      <td>
        <div class="rh-table-nome-cell">
          <span class="rh-avatar">{iniciais}</span>
          <span class="rh-table-nome-text">{row['Nome']}</span>
        </div>
      </td>
      <td>
        <div class="rh-table-alertas-cell">
          <span class="rh-alertas-num {sev_cls}">{qtd}</span>
          <span class="rh-badge {sev_cls}">{sev_lbl}</span>
        </div>
      </td>
      <td class="rh-table-comentario">{comentario}</td>
    </tr>"""

                    st.markdown(f"""
<div class="rh-table-wrap">
  <div class="rh-table-header">
    <div class="rh-table-header-title">⚠️ Alunos Desengajados</div>
    <div class="rh-table-header-count">{total_view} alunos encontrados</div>
  </div>
  <table class="rh-table">
    <thead>
      <tr>
        <th>Curso</th>
        <th>Turma</th>
        <th>Nome do Aluno</th>
        <th>Alertas</th>
        <th>Último Comentário</th>
      </tr>
    </thead>
    <tbody>{linhas_html}</tbody>
  </table>
</div>
""", unsafe_allow_html=True)

                    # ── Gerar Excel e enviar e-mail ──
                    excel_bytes = exportar_excel_bytes(df_rel)
                    data_hoje   = datetime.now().strftime('%d/%m/%Y')
                    dests = [email_usuario.strip()] if email_usuario and "@" in email_usuario else []
                    if SENDGRID_OK:
                        ok, msg = enviar_email(excel_bytes, dests, data_hoje, len(df_rel), criticos, atencao, monitorar)
                        if ok: st.success(f"📧 {msg}")
                        else:  st.warning(f"⚠️ E-mail não enviado: {msg}")
                    else:
                        st.warning("⚠️ Biblioteca SendGrid não instalada.")

                    # ── Botão de download ──
                    st.download_button(
                        "⚡ Baixar Relatório Excel",
                        data=excel_bytes,
                        file_name=f"relatorio_cs_{datetime.now().strftime('%d%m%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Erro ao processar: {e}")
                    st.info("Verifique se os arquivos corretos foram enviados com os filtros indicados.")

st.markdown('</div>', unsafe_allow_html=True)
st.markdown("""
<div class="rh-footer">
  <span>Rehagro</span><span class="rh-footer-dot"></span>
  <span>Customer Success</span><span class="rh-footer-dot"></span>
  <span>Agente de Engajamento</span><span class="rh-footer-dot"></span>
  <span>© 2026</span>
</div>
""", unsafe_allow_html=True)
